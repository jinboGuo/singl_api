import time
import requests
from openpyxl import load_workbook
from api_test_cases.get_execution_output_json import abs_dir
from basic_info.get_auth_token import get_headers
from new_api_cases.get_statementId import statementId_flow_use, preview_result_flow_use
from util.Open_DB import MYSQL
from basic_info.setting import MySQL_CONFIG
from basic_info.setting import host
from util.format_res import dict_res

host_for_url=host
# excel_dir=os.path.join(os.path.abspath(".."),r"api_test_cases\flow_dataset_info.xlsx")
table=abs_dir("flow_dataset_info.xlsx")
host=MySQL_CONFIG["HOST"]
port=MySQL_CONFIG["PORT"]
user=MySQL_CONFIG["USER"]
pwd=MySQL_CONFIG["PASSWORD"]
db=MySQL_CONFIG["DB"]


ms=MYSQL(host,user,pwd,db,port)
flow_table=load_workbook(table)
sheet=flow_table.get_sheet_by_name("flink")
max_rows=sheet.max_row
print(max_rows)


def get_flow_id():
    flow_id_list=[]
    for i in range(2,max_rows+1):
        flow_id_list.append(sheet.cell(row=i,column=2).value)
    return flow_id_list

def data_for_exe():
    data_for_exe_list=[]
    for i in get_flow_id():
        try:
            sql='select name from merce_flow where id="%s"'%i
            result=ms.ExecuQuery(sql)
            flow_name=result[0]["name"]
        except Exception as e:
            print("flow_name不存在{}".format(e))
            return
        data={
            "creator":"admin",
            "flowId":i,
            "flowVersion":2,
            "flowName":flow_name,
            "flowType":"rtcflow",
            "properties":{
                "engine":"flink",
                "debug":False,
                "runtimeSettings":{
                    "master":"yarn",
                    "queue":"default",
                    "nodeLabel":"",
                    "driverMemory":2048,
                    "executorMemory":2048,
                    "executorCores":1,
                    "parallelism":1,
                    "useLatestState":False,
                    "allowNonRestoredState":False,
                    "lineageEnable":False,
                    "savepointDir":"",
                    "kerberosEnable":False,
                    "kerberosPrincipal":"",
                    "kerberosKeytab":"",
                    "kerberosJaasConf":""
                },
                "checkpointSettings":{
                    "checkpointEnable":True,
                    "checkpointMode":"exactly_once",
                    "checkpointDir":"hdfs:///tmp/flink/checkpoints/info4",
                    "checkpointAsync":True,
                    "checkpointStateBackend":"rocksdb",
                    "checkpointIncremental":False,
                    "checkpointInterval":60000,
                    "checkpointMinpause":5000,
                    "checkpointTimeout":600000,
                    "checkpointExternalSave":True
                },
                "restartStrategySettings":{
                    "restartStrategy":"FixedDelayRestart",
                    "restartMaxAttempts":3,
                    "restartInterval":60,
                    "restartDelayInterval":10
                },
                "latencyTrackingSettings":{
                    "latencyTrackingEnable":False,
                    "latencyTrackingInterval":60000
                }
            }}
        data_for_exe_list.append(data)
    return data_for_exe_list

def create_flink_exe():
    flink_exe_id=[]
    for i in data_for_exe():
        result=requests.post(url=host_for_url+"/api/executions/rtcflow",json=i,headers=get_headers(host_for_url))
        print(host_for_url+"/api/executions/rtcflow",get_headers(host_for_url))
        print(result.status_code,result.text)
        flink_exe_id.append(dict_res(result.text)["id"])
    return flink_exe_id

def get_exe_result():
    sink_result_list= []
    sink_result_dict={}
    for i in create_flink_exe():
        try:
            sql="select flow_id,flow_type,status_type from merce_flow_execution where id='%s'"%i
            result=ms.ExecuQuery(sql)
            sink_result_dict["flow_id"]=result[0]["flow_id"]
            sink_result_dict["status_type"]=result[0]["status_type"]
            e=get_e_finial_status(i,sink_result_dict["status_type"])
            sink_result_dict["status_type"]=e
            g_result=requests.get(url=host_for_url+"/api/executions/rtc/"+i+"/output",headers=get_headers(host_for_url))
            sink_result_dict["dataset_id"]=dict_res(g_result.text)[0]["id"]
            sink_result_list.append(sink_result_dict)
        except:
            return
    return sink_result_list

def get_e_finial_status(id,status_type):
    while 1:
        if status_type == "READY":
            time.sleep(20)
            try:
                sql="select status_type from merce_flow_execution where id='%s'"%id
                result=ms.ExecuQuery(sql)
                status_type=result[0]["status_type"]
            except:
                return
        elif status_type in ("RUNNING","FAILED","KILLED"):
            return status_type
        else:
            return



def write_act_result():
    for i in get_exe_result():
        dataset_id=i['dataset_id']
        flow_id=i['flow_id']
        status_type=i['status_type']
        if status_type =="RUNNING":
            statementID = statementId_flow_use(host_for_url, dataset_id)
            result=preview_result_flow_use(host_for_url, dataset_id, statementID)
            while result==0:
                statementID = statementId_flow_use(host_for_url, dataset_id)
                result=preview_result_flow_use(host_for_url, dataset_id, statementID)
            for i in range(2,sheet.max_row+1):
                if sheet.cell(row=i,column=4).value==dataset_id:
                    sheet.cell(row=i,column=8,value=str(result))
                    sheet.cell(row=i,column=6,value=status_type)
                else:
                    print("请检查数据集id是否正确")
        elif status_type in ("FAILED", "KILLED"):
            for i in range(2,sheet.max_row+1):
                if sheet.cell(row=i,column=4).value==dataset_id:
                    sheet.cell(row=i,column=8,value="")
                    sheet.cell(row=i,column=6,value=status_type)
        else:
            print("请检查status_type状态是否正确")

    flow_table.save(table)


def check_reult():
    write_act_result()
    for i in range(2,sheet.max_row+1):
        exp=sheet.cell(row=i,column=7).value
        ee=list(eval(exp))
        res=sheet.cell(row=i,column=8).value
        rr=list(eval(res))
        e_keys=[]
        for key in ee[0].keys():
            e_keys.append(key)
        expect_len=len(ee)
        rr1=rr[-expect_len:] #实际结果切片和预期结果长度一致的数据结果按key排序之后，有一个字段相等认为相等
        result = []
        for t in e_keys:
            a=sorted(ee,key=lambda i:i[t])
            b=sorted(rr1,key=lambda i:i[t])
            result.append(a == b)
        if True in result:
            sheet.cell(row=i, column=9, value="pass")
            print('test_result:', sheet.cell(row=i, column=9).value)
            sheet.cell(row=i, column=10, value="")
        else:
            sheet.cell(row=i, column=9, value="fail")
            sheet.cell(row=i, column=10, value="")
    flow_table.save(table)


if __name__=="__main__":
    check_reult()


