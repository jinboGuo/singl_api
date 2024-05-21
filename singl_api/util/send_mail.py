# coding:utf-8
import time, os
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
import sys
from smtplib import SMTP_SSL
from openpyxl import load_workbook
from basic_info.setting import result_path, email_user, host_server, pwd ,log

current_path = os.path.abspath(os.path.dirname(__file__))
root_path = os.path.split(current_path)[0]
sys.path.append(root_path)

def baymax_main(cases_dir,host, receivers, sheet, start_time,total_time):
    msg = MIMEMultipart()
    """邮件的正文内容----API执行结果,统计api执行结果，加入到邮件正文中，失败的用例name：失败的原因"""
    api_cases_table = load_workbook(cases_dir)
    cases_sheet = api_cases_table.get_sheet_by_name(sheet)
    sheet_rows = cases_sheet.max_row
    cases_num = sheet_rows - 1
    non_cases = 0
    pass_cases = 0
    failed_cases = 0
    failed_cases_detail = {}
    failed_cases_name = []
    for row in range(2, sheet_rows+1):
        if cases_sheet.cell(row=row, column=14).value == 'pass':
            pass_cases += 1
        elif cases_sheet.cell(row=row, column=14).value == 'fail':
            failed_cases += 1
            case_name = cases_sheet.cell(row=row, column=2).value
            failed_cases_detail[case_name] = cases_sheet.cell(row=row, column=15).value   # 将用例name和失败原因列出
            failed_cases_name.append(cases_sheet.cell(row=row, column=2).value)
        else:
            non_cases += 1

    """邮件的正文内容"""
    if failed_cases != 0:
        mail_content = """各位好:
        本次用例执行环境：%s
        开始时间: %s
        合计耗时: %s
        用例总数: %d 条
        用例成功: %d 条
        用例失败: %d 条
        用例跳过: %d 条
        通过率: %s
        用例执行详情请查看附件报告！
        失败用例名称如下：%s
        """ % (host,start_time,total_time, cases_num, pass_cases, failed_cases, non_cases,str(float(pass_cases+non_cases)/cases_num*100)+'%', failed_cases_name)
    else:
        mail_content = """各位好:
        本次用例执行环境：%s
        开始时间: %s
        合计耗时: %s
        用例总数: %d 条
        用例成功: %d 条
        用例跳过: %d 条
        通过率: %s
        用例执行详情请查看附件报告！""" % (host, start_time,total_time,cases_num, pass_cases, non_cases,str(float(pass_cases+non_cases)/cases_num*100)+'%')
    """ 邮件标题"""
    mail_title = time.strftime("%Y-%m-%d", time.localtime()) + "号-" + sheet +'环境API自动化测试报告'

    body = MIMEText(mail_content,"plain", "utf-8")
    msg["Subject"] = Header(mail_title, "utf-8")
    msg["From"] = email_user
    if len(receivers)>1:
        msg["To"] = ",".join(receivers)
    else:
        msg["To"] = receivers[0]
    msg.attach(body)

    """附件:附件名称用英文"""
    att = open(result_path, "r", encoding="utf-8").read()
    att = MIMEText(att, "base64", "utf-8")
    att["Content-Type"] = "application/octet-stream"
    att["Content-Disposition"] = "attachment;filename='%s'" % result_path
    msg.attach(att)

    """ssl登录"""
    smtp = SMTP_SSL(host_server)
    smtp.set_debuglevel(0)
    smtp.ehlo(host_server)
    smtp.login(email_user, pwd)
    if len(receivers)>1:
        smtp.sendmail(email_user, ",".join(receivers), msg.as_string())
        log.info('%s----发送邮件成功' % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        smtp.quit()
    else:
        smtp.sendmail(email_user, receivers[0], msg.as_string())
        log.info('%s----发送邮件成功' % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        smtp.quit()