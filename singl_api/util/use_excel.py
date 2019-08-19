import xlrd, xlwt
from openpyxl import workbook
from openpyxl import load_workbook
from xlutils.copy import copy
import os
abs_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))

# 设置表格样式
def set_style(name, height, bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    return style


# 写Excel
def write_excel():
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('flow_info', cell_overwrite_ok=True)
    row0 = ["id", "flow_id", "except_result", "actual", "test_result"]
    colum0 = ["35033c8d-fadc-4628-abf9-6803953fba34"]
    # 写第一行
    for i in range(0, len(row0)):
        sheet1.write(0, i, row0[i], set_style('Times New Roman', 220, True))
    # 写第二行的第二列
    for i in range(0, len(colum0)):
        sheet1.write(i + 1, 0, i + 1, set_style("Times New Roman", 220, True))
        sheet1.write(i + 1, 1, colum0[i], set_style('Times New Roman', 220, True))

    sheet1.write(1, 2, 'data_json')
    f.save('flow_dataset_info.xls')


# 读
def read_execl():
    data = xlrd.open_workbook("flow_dataset_info.xls")
    table1 = data.sheets()[0]
    table2 = data.sheet_by_name("flow_info")
    data1 = table1.ncols
    data2 = table1.nrows
    data3 = table1.cell(0, 2).value
    print(table1)
    return data1, data2, data3


# 改
def update_excel():
    wb = xlrd.open_workbook("flow_dataset_info.xls")  # 打开表
    crb = copy(wb)  # 复制表
    crb_sheet = crb.get_sheet(0)  # 打开复制表的第一个表单
    # crb_sheet.write(1, 3, "") # 第2行第四列，修改为“”

    crb_sheet.write(1, 3, )
    crb.save("flow_dataset_info.xls")


def check():
    table = xlrd.open_workbook("flow_dataset_info.xls")
    table_sheet = table.sheets()[0]
    c_table = copy(table)
    c_table_sheet = c_table.get_sheet(0)

    c_rows = table_sheet.nrows

    print('行数：', c_rows)
    # print(table_sheet.cell(1, 2).value)
    for i in range(1, c_rows):
        # print(table_sheet.cell(i, 2).value)
        if table_sheet.cell(i, 2).value:
            if table_sheet.cell(i, 2).value == table_sheet.cell(i, 3).value:
                c_table_sheet.write(i, 4, "pass")
            else:
                c_table_sheet.write(i, 4, "fail")

    c_table.save("flow_dataset_info.xls")
def openpyxl_read():
    wb = load_workbook('test.xlsx')
    sheetnames = wb.get_sheet_names()
    flow_info = wb.get_sheet_by_name(sheetnames[0])
    rows = flow_info.max_row
    columns = flow_info.max_column
    flow_info.cell(row=2,column=12,value=8888)
    print(flow_info.cell(row=1, column=3).value, rows, columns)
    print(flow_info.cell(row=2, column=12).value)


def get_value():
    table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
    table_sheet = table.get_sheet_by_name('flow_info')
    c_rows = table_sheet.max_row

    # mode = overwrite:实际结果写入表后，对比预期结果和实际结果,并把失败详情存在 fail_detail
    print('-----开始对比结果----')


    va7 = list(eval(table_sheet.cell(row=11, column=7).value))
    va8 = list(eval(table_sheet.cell(row=11, column=8).value))
    print(va7)
    print(va8)
    S_va7 = sorted(va7, key=lambda item: item["id"], reverse=True)
    S_va8 = sorted(va8, key=lambda item: item["id"], reverse=True)
    print(S_va7==S_va8)
    print(S_va8)


    # s_va7 = sorted(va7, key=lambda item: item[-1], reverse=True)
    # print(s_va7)

    # for i in range(2, c_rows+1):
    #     table_sheet.cell(row=i, column=1, value=i-1)
    #     if table_sheet.cell(row=i, column=11).value == 'overwrite':  # 判断mode
    # # 实际结果存在
    #         if table_sheet.cell(row=i, column=8).value and table_sheet.cell(row=i, column=6).value == "SUCCEEDED":
    #             # 实际结果和预期结果相等
    #             if table_sheet.cell(row=i, column=7).value == table_sheet.cell(row=i, column=8).value:
    #                 table_sheet.cell(row=i, column=9, value="pass")
    #                 print('test_result:', table_sheet.cell(row=i, column=9).value)
    #                 table_sheet.cell(row=i, column=10, value="")

if __name__ == '__main__':
    print(get_value())
