# coding:utf-8
import json
import os
import re
import time
import random
from openpyxl.styles import PatternFill, colors
from new_api_cases.get_statementId import get_sql_analyse_dataset_info, get_sql_analyse_statement_id, get_sql_execte_statement_id
from util import myddt
import xlrd
from openpyxl import load_workbook
import requests
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from util.format_res import dict_res
from basic_info.setting import dw_host, dw_sheet, dw_cases_dir, log, resource_type
from basic_info.get_auth_token import get_headers_root, get_headers
from new_api_cases.dw_deal_parameters import deal_parameters
import unittest
from new_api_cases.dw_prepare_datas import get_asset_directory, update_asset_directory, move_asset_directory, \
    duplicate_asset_directory, duplicate_move_asset_directory, delete_asset_directory,update_asset,create_sql_asset, \
    sql_analyse_data, batch_create_asset, get_improt_data
from util.get_deal_parameter import get_resourceid

cases_dir = dw_cases_dir
case_table = load_workbook(cases_dir)
dw_master=dw_sheet
case_table_sheet = case_table.get_sheet_by_name(dw_master)
all_rows = case_table_sheet.max_row
host = dw_host
woven_dir = os.path.join(os.path.abspath('.'),'attachment\\import_auto_apitest_df.woven').replace('\\', '/')


def deal_request_method():
    """
    判断请求方法，并根据不同的请求方法调用不同的处理方式
    :return:
    """
    try:
        for i in range(2, all_rows+1):
            request_method = case_table_sheet.cell(row=i, column=4).value
            request_method_upper = request_method.upper()
            request_url = host+case_table_sheet.cell(row=i, column=5).value
            old_data = case_table_sheet.cell(row=i, column=6).value
            request_data = deal_parameters(old_data,request_method_upper,request_url)
            log.info("request  data：%s" % request_data)
            api_name = case_table_sheet.cell(row=i, column=1).value
            is_run = case_table_sheet.cell(row=i, column=16).value
            if request_method_upper:
                if is_run =='Y' or is_run=='y':
                    if api_name == 'tenants':
                        """
                        租户的用例需要使用root用户登录后操作
                        根据不同的请求方法，进行分发
                        """
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers_root())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers_root())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                    else:
                        """根据不同的请求方法，进行分发"""
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                else:
                    log.info(" 第%d 行脚本未执行，请查看isRun是否为Y或者y！"% i)
            else:
                log.info("第 %d 行请求方法为空" % i)
        '''执行结束后保存表格'''
        case_table.save(cases_dir)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(i, e))



def post_request_result_check(row, column, url, headers, data, table_sheet_name):
    """
    POST接口请求，脚本里post请求的处理
    :param row:
    :param column:
    :param url:
    :param headers:
    :param data:
    :param table_sheet_name:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s " % case_detail)
        log.info("请求url：%s" % url)
        if case_detail == '添加资产目录':
            new_data = get_asset_directory(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '资产目录移动':
            parent_id, new_data = move_asset_directory(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            new_url = url.format(parent_id)
            log.info("new_url：%s " % new_url)
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '添加重名资产目录':
            new_data = duplicate_asset_directory(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '创建服务视图资产' in case_detail:
            new_data = create_sql_asset(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '初始化Sql' in case_detail:
            new_data = sql_analyse_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '获取SQL执行任务结果':
            exec_use_params = get_sql_analyse_dataset_info(host, data)
            exec_use_params = json.dumps(exec_use_params, separators=(',', ':'))
            statement_id, session_id, cluster_id = get_sql_execte_statement_id(host, data)
            new_url = url.format(statement_id, session_id, cluster_id)
            log.info("new_url：%s " % new_url)
            response = requests.post(url=new_url, headers=headers, data=exec_use_params)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            while "waiting" in response.text or "running" in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=new_url, headers=headers, data=exec_use_params)
                time.sleep(5)
                count_num += 1
                if count_num == 50:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '批量创建数据集资产':
            new_data = batch_create_asset(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "创建模型woven导入任务" == case_detail:
                files = {"file": ('import_auto_apitest_df.woven', open(woven_dir, 'rb'), "application/octet-stream"),
                         "name": (None, "gjb_type_df_import" + str(random.randint(0, 999999)), None),
                         "remark": (None, "gjb_type_df_import", None), "taskType": (None, "offlineDev", None),
                         "flowResourceId": (None, str(get_resourceid(resource_type[3])), None),
                         "flowResourcePath": (None, "Flows;", None),
                         "datasetResourceId": (None, str(get_resourceid(resource_type[1])), None),
                         "datasetResourcePath": (None, "Datasets;", None),
                         "datasourceResourceId": (None, str(get_resourceid(resource_type[0])), None),
                         "datasourceResourcePath": (None, "Datasources;", None),
                         "schemaResourceId": (None, str(get_resourceid(resource_type[2])), None),
                         "schemaResourcePath": (None, "Schemas;", None)}
                headers.pop('Content-Type')
                response = requests.post(url=url, files=files, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型woven文件" == case_detail:
                flow_import_task_info_id, new_data = get_improt_data(data)
                new_url = url.format(flow_import_task_info_id)
                log.info("请求url：%s" % new_url)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.post(url=new_url, headers=get_headers(), data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4,
                             value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "创建元模型" == case_detail:
            log.info("request data： %s" % data)
            response = requests.post(url=url, headers=get_headers(), data=data.encode('utf-8'))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        else:
            if data:
                data = str(data)
                if "{}" in url:
                    data = data.split("##")
                    new_url = url.format(data[1])
                    log.info("请求new_url：%s" % new_url)
                    response = requests.post(url=new_url, headers=headers, data=data[0].encode('utf-8'))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif data.startswith('{') and data.endswith('}'):
                    data_dict = dict_res(data)
                    response = requests.post(url=url, headers=headers, json=data_dict)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif data.startswith('[') and data.endswith(']'):
                    response = requests.post(url=url, headers=headers, data=data)
                    time.sleep(3)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    log.error("请求体数据错误！")
            else:
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(1)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例{}执行过程中出错{}".format(case_detail, e))



def get_request_result_check(url, headers, data, table_sheet_name, row, column):
    """
    GET接口请求需要从parameter中获取参数,并把参数拼装到URL中
    :param url:
    :param headers:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s " %case_detail)
        log.info("请求url：%s" % url)
        if data:
            if case_detail == '根据statement id,获取Sql Analyze结果(获取输出字段)':
                statement_id, session_id, cluster_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(statement_id, session_id, cluster_id)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 0
                while "waiting" in response.text or "running" in response.text:
                    log.info("再次查询前：%s %s" % (response.status_code, response.text))
                    response = requests.get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '结束指定statementId对应的查询任务':  # 取消SQL analyse接口
                cancel_statement_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(cancel_statement_id)
                log.info("new_url：%s " % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                new_url = url.format(data)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            response = requests.get(url=url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


def put_request_result_check(url, row, data, table_sheet_name, column, headers):
    """
    PUT接口请求
                if '&' in str(data):
                # 分隔参数
                parameters = data.split('&')
                # 拼接URL
                new_url = url.format(parameters[0])
    :param url:
    :param row:
    :param data:
    :param table_sheet_name:
    :param column:
    :param headers:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s " %case_detail)
        log.info("请求url：%s" % url)
        if data :
            if case_detail == '资产目录重命名':
                new_data = update_asset_directory(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '资产目录重命名-名称重复':
                new_data = duplicate_move_asset_directory(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '更新资产' in case_detail:
                asset_id, new_data = update_asset(data)
                url = url.format(asset_id)
                log.info("new_url：%s " % url)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '&&' in str(data):
                '''分隔参数'''
                parameters = data.split('&&')
                '''拼接URL'''
                new_url = url.format(parameters[0])
                log.info("new_url：%s" % new_url)
                '''发送的参数体'''
                parameters_data = parameters[1]
                if parameters_data.startswith('{'):
                    response = requests.put(url=new_url, headers=headers, json=dict_res(parameters_data))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(table_sheet_name, row, column, response.status_code)
                    write_result(table_sheet_name, row, column + 4, response.text)
                else:
                    log.info('请确认第%d行parameters中需要update的值格式，应为id&{data}' % row)
            elif "{}" in url:
                new_url = url.format(data["id"])
                log.info("new_url：%s" % new_url)
                response = requests.put(url=new_url, headers=headers, data=json.dumps(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
            elif data.startswith('{') and data.endswith('}'):
                response = requests.put(url=url, headers=headers, data=data.encode('utf-8'))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
            else:
                response = requests.put(url=url, headers=headers, json=dict_res(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
    except Exception as e:
            log.error("{}执行过程中出错{}".format(case_detail, e))
        
def delete_request_result_check(url, data, table_sheet_name, row, column, headers):
    """
    delete接口请求
    :param url:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :param headers:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s " %case_detail)
        log.info("请求url：%s" % url)
        if data:
            if "删除资产目录" == case_detail:
                asset_id, new_data = delete_asset_directory(data)
                log.info("request   data：%s " % new_data)
                new_url = url.format(asset_id)
                log.info("new_url：%s " % new_url)
                response = requests.delete(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif isinstance(data, str):
                log.info("data：%s" % data)
                if "{}" in url:
                    new_url = url.format(data)
                    log.info("new_url：%s" % new_url)
                    response = requests.delete(url=new_url, headers=headers)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    response = requests.delete(url=url, headers=headers, data=json.dumps(data))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(table_sheet_name, row, column, response.status_code)
                    write_result(table_sheet_name, row, column + 4, response.text)
            elif isinstance(data, list):
                response = requests.delete(url=url, headers=headers, data=json.dumps(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                log.error("{}执行过程中出错".format(case_detail))
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))



def write_result(sheet, row, column, value):
    """
    写入返回结果
    :param sheet:
    :param row:
    :param column:
    :param value:
    :return:
    """
    sheet.cell(row=row, column=column, value=value)



def clean_vaule(sheet, row, column):
    """
    写入结果前，先把结果和对比结果全部清空
    :param sheet:
    :param row:
    :param column:
    :return:
    """
    sheet.cell(row=row, column=column, value='')
    sheet.cell(row=row, column=column+1, value='')
    sheet.cell(row=row, column=column + 4, value='')
    sheet.cell(row=row, column=column + 5, value='')
    sheet.cell(row=row, column=column + 6, value='')
    sheet.cell(row=row, column=column + 7, value='')



def read_data():
        data = xlrd.open_workbook(cases_dir)
        table = data.sheet_by_name(dw_master)
        """获取总行数"""
        nrows = table.nrows
        if nrows > 1:
            """获取第一行的内容，列表格式"""
            keys = table.row_values(0)
            list_api_data = []
            """获取每一行的内容，列表格式"""
            for col in range(1, nrows):
                values = table.row_values(col)
                """ keys，values组合转换为字典"""
                api_dict = dict(zip(keys, values))
                if api_dict['is_run']=="y" or api_dict['is_run']=="Y":
                    list_api_data.append(api_dict)
            return list_api_data
        else:
            log.info("表格是空数据!")
            return None

testdata = read_data()



@myddt.ddt
class CheckResult(unittest.TestCase):
    def compare_code_result(self):
        """1.对比预期code和接口响应返回的status code"""
        for row in range(2, all_rows+1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            """预期status code和接口返回status code"""
            ex_status_code = case_table_sheet.cell(row=row, column=7).value
            ac_status_code = case_table_sheet.cell(row=row, column=8).value
            """判断两个status code是否相等"""
            if is_run == 'Y' or is_run == 'y':
                if ex_status_code and ac_status_code != '':
                    if ex_status_code == ac_status_code:
                        case_table_sheet.cell(row=row, column=9, value='pass').fill = PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        case_table_sheet.cell(row=row, column=9, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=13, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                        (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("第 %d 行 status_code为空" %row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    def compare_text_result(self):
        """对比预期response和实际返回的response.text，根据预期和实际结果的关系进行处理"""
        for row in range(2, all_rows+1):
            """接口返回的response.text"""
            response_text = case_table_sheet.cell(row=row, column=12).value
            response_text_dict = dict_res(response_text)
            """预期结果"""
            expect_text = case_table_sheet.cell(row=row, column=10).value
            """接口关键字"""
            key_word = case_table_sheet.cell(row=row, column=3).value
            """status_code对比结果"""
            code_result = case_table_sheet.cell(row=row, column=9).value
            """预期text和response.text的关系"""
            relation = case_table_sheet.cell(row=row, column=11).value
            is_run = case_table_sheet.cell(row=row, column=16).value
            if is_run == 'Y' or is_run == 'y':
                """
                1.status_code 对比结果pass的前提下，判断response.text断言是否正确,
                2.status_code 对比结果fail时，用例整体结果设为fail
                """
                if code_result == 'pass':
                    if key_word in ('create', 'query', 'update', 'delete'):
                        self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                    else:
                        log.error("请确认第%d行的key_word" % row)
                elif code_result == 'fail':
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("请确认第 %d 行 status_code对比结果" % row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)

        case_table.save(cases_dir)


    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        """
        :param key_word:
        :param relation:
        :param expect_text:
        :param response_text:
        :param response_text_dict:
        :param row:
        :param column:
        :return: 根据expect_text, response_text的关系，进行断言, 目前只处理了等于和包含两种关系
        """
        if key_word == 'create':
            if relation == '=':
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']), '第%d行的response_text长度和预期不一致' % row)
                        except:
                            #log.info("第 %d 行 response_text返回的id和预期id长度不一致" %row)
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '第%d行的expect_text:%s和response_text:%s不相等' % (row,expect_text, response_text))
                        except:
                            #log.info("第%d行的expect_text:%s和response_text:%s不相等" %(row,expect_text, response_text))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    '''只返回一个id串的情况下，判断预期长度和id长度一致'''
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行的response_text长度和预期不一致' % row)
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                """返回多内容时，断言多个值可以用&连接，并且expect_text包含在response_text中"""
                if expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relatrion" %row)
                case_table_sheet.cell(row=row, column=column, value='请确认%d行 的预期text和接口response.text的relatrion'%row)
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}', '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行expect_text和response_text不相等' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertEqual(expect_text, response_text, '第%d行expect_text:%s和response_text:%s不相等' % (row,expect_text,response_text))
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                if expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relation" % row)
                case_table_sheet.cell(row=row, column=column, value='请确认第 %d 行 预期expect_text和response_text的relation'%row)
        else:
            log.error("请确认第 %d 行 的key_word" % row)
        case_table.save(cases_dir)


    def deal_result(self):
        """
        对比code
        self.compare_code_result()
        对比text
        self.compare_text_result()
        根据code result和text result判断case最终结果
        :return: 对比case最终的结果
        """
        self.compare_code_result()
        self.compare_text_result()
        for row in range(2, all_rows + 1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            status_code_result = case_table_sheet.cell(row=row, column=9).value
            response_text_result = case_table_sheet.cell(row=row, column=13).value
            if is_run=='Y' or is_run=='y':
                if status_code_result == 'pass' and response_text_result == 'pass':
                    #log.info("测试用例-%s pass" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    case_table_sheet.cell(row=row, column=15, value='')
                elif status_code_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code对比失败,预期为%s,实际为%s' \
                                                                    % (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                elif status_code_result == 'pass' and response_text_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回内容对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=10).value, case_table_sheet.cell(row=row, column=12).value))
                else:
                    log.error("请确认status code或response.text对比结果")
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    @myddt.data(*testdata)
    def test_api(self,data):
        self.case_name = data['case_detail']
        self.url=host+data['url']
        self.method=data['method']
        self.case_result = data['case_result']
        self.result2 = data['result2']
        self.header=get_headers()
        self.body=data['parameters']
        self.expect_text = data['expect_text']
        self.extract_data=data['response_text']
        self.readData_code =data["response__status_code"]
        print("******* 执行用例 ->{0} *********".format(self.case_name))
        print("请求URL: {0}".format(self.url))
        print("请求方式: {0}".format(self.method))
        print("请求header:{0}".format(self.header))
        print("请求body:{0}".format(self.body))
        if self.case_result == 'pass':
            print("返回状态码：%s 响应信息：%s" % (self.readData_code,self.extract_data))
            self.assertIn(self.expect_text,self.extract_data,"返回实际结果是->:%s" % self.extract_data)
        else:
            print("返回状态码：%s 响应信息：%s" % (self.readData_code, self.extract_data))
            self.assertIn(self.expect_text, self.extract_data, "返回实际结果是->:%s" % self.extract_data)