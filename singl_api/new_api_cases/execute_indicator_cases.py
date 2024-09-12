# coding:utf-8
import json
import os
import re
from openpyxl.styles import PatternFill, colors
from util import myddt
from openpyxl import load_workbook
import requests
from util.format_res import dict_res
from basic_info.setting import indicator_sheet, indicator_cases_dir, log
from basic_info.get_auth_token import get_headers, get_headers_root
from new_api_cases.indicator_deal_parameters import deal_parameters
import unittest
import time
from basic_info.setting import indicator_host
from util.get_deal_parameter import get_indicator_dimdir_id, get_current_time, get_indicator_id, \
    get_indicator_subtask_id, get_indicator_dir_id,indicator_data

cases_dir = indicator_cases_dir
case_table = load_workbook(cases_dir)
indicator_master = indicator_sheet
case_table_sheet = case_table.get_sheet_by_name(indicator_master)
all_rows = case_table_sheet.max_row
host = indicator_host


def deal_request_method():
    """
    判断请求方法，并根据不同的请求方法调用不同的处理方式
    :return:
    """
    try:
        for i in range(2, all_rows + 1):
            request_method = case_table_sheet.cell(row=i, column=4).value
            request_method_upper = request_method.upper()
            request_url = host + case_table_sheet.cell(row=i, column=5).value
            old_data = case_table_sheet.cell(row=i, column=6).value
            request_data = deal_parameters(old_data, request_method_upper, request_url)
            log.info("request  data：%s" % request_data)
            api_name = case_table_sheet.cell(row=i, column=1).value
            is_run = case_table_sheet.cell(row=i, column=16).value
            if request_method_upper:
                if is_run == 'Y' or is_run == 'y':
                    if api_name == 'tenants':
                        """
                        租户的用例需要使用root用户登录后操作
                        根据不同的请求方法，进行分发
                        """
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers_root(),
                                                      data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers_root(), data=request_data,
                                                     table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data,
                                                     table_sheet_name=case_table_sheet, column=8,
                                                     headers=get_headers_root())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data,
                                                        table_sheet_name=case_table_sheet, row=i, column=8,
                                                        headers=get_headers_root())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                    else:
                        """根据不同的请求方法，进行分发"""
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers(),
                                                      data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers(), data=request_data,
                                                     table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data,
                                                     table_sheet_name=case_table_sheet, column=8, headers=get_headers())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data,
                                                        table_sheet_name=case_table_sheet, row=i, column=8,
                                                        headers=get_headers())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                else:
                    log.info(" 第%d 行脚本未执行，请查看isRun是否为Y或者y！" % i)
            else:
                log.info("第 %d 行请求方法为空" % i)
        '''执行结束后保存表格'''
        case_table.save(cases_dir)
    except Exception as e:
        log.error("执行过程中出错{}".format( e))


# POST请求
def post_request_result_check(row, column, url, headers, data, table_sheet_name):
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        if case_detail == '上传驱动包-mysql':
            driver_file_path = os.path.join(os.path.abspath('.'), 'attachment\\mysql-connector-j-8.0.32_driver.jar').replace('\\','/')
            files = {
                'file': (os.path.basename(driver_file_path), open(driver_file_path, 'rb')),
                'dbType': (None, 'Mysql'),
            }
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传驱动包-hive':
            driver_file_path = os.path.join(os.path.abspath('.'), 'attachment\\hive-jdbc-3.1.0.3.1.5.0-152_driver.jar').replace('\\','/')
            files = {
                'file': (os.path.basename(driver_file_path), open(driver_file_path, 'rb')),
                'dbType': (None, 'HIVE'),
            }
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传驱动包-dm':
            driver_file_path = os.path.join(os.path.abspath('.'), 'attachment\\dm7-18.0.0_driver.jar').replace('\\','/')
            files = {
                'file': (os.path.basename(driver_file_path), open(driver_file_path, 'rb')),
                'dbType': (None, 'Dm'),
            }
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传驱动包-Oushu':
            driver_file_path = os.path.join(os.path.abspath('.'), 'attachment\\postgresql-42.2.23_driver.jar').replace('\\','/')
            files = {
                'file': (os.path.basename(driver_file_path), open(driver_file_path, 'rb')),
                'dbType': (None, 'OushuDB'),
            }
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '创建数据库驱动' in case_detail:
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '任务执行':
            url = url.format(get_indicator_id())
            log.info("request   url：%s" % url)
            headers.pop('Content-Type')
            response = requests.post(url=url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '子任务重跑':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '子任务预览':
            url = url.format(get_indicator_subtask_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '子任务预览':
            url = url.format(get_indicator_subtask_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '结果查询':
            url = url.format(get_indicator_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '下载指标结果':
            url = url.format(get_indicator_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            try:
                response_content = response.content.decode('utf-8')
                if response_content:
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value='False')
            except UnicodeDecodeError:
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value='True')
        elif case_detail == '指标更换目录':
            url = url.format(get_indicator_dir_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == 'sql试运行':
            log.info("request   url：%s" % url)
            headers['Content-Type'] = 'application/x-www-form-urlencoded'
            data = indicator_data()
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '删除数据集':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '删除数据源-删除自动化创建的':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '删除数据库驱动':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '删除驱动包':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '删除元数据':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '查找相同维度的指标':
            url = url.format(get_indicator_id())
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '复合指标试运行':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data.encode('utf-8'))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建复合指标':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data.encode('utf-8'))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '复合指标任务下线':
            log.info("request   url：%s" % url)
            response = requests.post(url=url, headers=headers, data=data.encode('utf-8'))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            if data:
                data = str(data)
                if data.startswith('{') and data.endswith('}'):
                    data_dict = dict_res(data)
                    response = requests.post(url=url, headers=headers, json=data_dict)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif data.startswith('[') and data.endswith(']'):
                    response = requests.post(url=url, headers=headers, data=data)
                    time.sleep(3)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    log.error("请求体数据错误！")
            else:
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(1)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例{}执行过程中出错{}".format(case_detail, e))


# GET请求
def get_request_result_check(url, headers, data, table_sheet_name, row, column):
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        # GET请求需要从parameter中获取参数,并把参数拼装到URL中，
        if data:
            if case_detail == '子任务查询日志':
                new_url = url.format(get_indicator_id())
                log.info("request   url：%s" % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '结果查询信息':
                url = url.format(get_indicator_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '获取查询维度信息':
                url = url.format(get_indicator_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '获取指标信息':
                url = url.format(get_indicator_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '获取目录信息':
                url = url.format(get_indicator_dir_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查询指标总数':
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查询各目录中指标数':
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查询指定时间段内的指标数':
                url = url.format(get_current_time()[:10], get_current_time()[:10])
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查看指标目录详情':
                url = url.format(get_indicator_dir_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查询血缘':
                url = url.format(get_indicator_id())
                log.info("request   url：%s" % url)
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


# PUT请求
def put_request_result_check(url, row, data, table_sheet_name, column, headers):
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        if data:
            if case_detail == '修改指标目录':
                new_url = url.format(get_indicator_dir_id())
                log.info("request   url：%s" % url)
                response = requests.put(url=new_url, headers=headers, data=data.encode('UTF-8'))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '维度目录重命名':
                new_url = url.format(get_indicator_dimdir_id())
                log.info("request   url：%s" % url)
                response = requests.put(url=new_url, headers=headers, data=data.encode('UTF-8'))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


def delete_request_result_check(url, data, table_sheet_name, row, column, headers):
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        if isinstance(data, str):
            log.info("data：%s" % data)
            if "{}" in url:
                new_url = url.format(data)
                log.info("new_url：%s" % new_url)
                response = requests.delete(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                response = requests.delete(url=url, headers=headers, data=data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
        elif isinstance(data, list):
            response = requests.delete(url=url, headers=headers, data=json.dumps(data))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)

        else:
            log.error("{}执行过程中出错".format(case_detail))
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


#  写入返回结果
def write_result(sheet, row, column, value):
    sheet.cell(row=row, column=column, value=value)


#  写入结果前，先把结果和对比结果全部清空
def clean_vaule(sheet, row, column):
    sheet.cell(row=row, column=column, value='')
    sheet.cell(row=row, column=column + 1, value='')
    sheet.cell(row=row, column=column + 4, value='')
    sheet.cell(row=row, column=column + 5, value='')
    sheet.cell(row=row, column=column + 6, value='')
    sheet.cell(row=row, column=column + 7, value='')


def read_data():
    data = load_workbook(cases_dir)
    table = data[indicator_master]
    """获取总行数"""
    nrows = sum(1 for _ in table.iter_rows())
    if nrows > 1:
        """获取第一行的内容，列表格式"""
        keys = [cell.value for cell in table[1]]
        list_api_data = []
        """获取每一行的内容，列表格式"""
        for row in range(2, nrows + 1):  # 从第二行开始，因为第一行被认为是标题行
            values = [cell.value for cell in table[row]]
            api_dict = dict(zip(keys, values))
            """ keys，values组合转换为字典"""
            if api_dict['is_run'] == "y" or api_dict['is_run'] == "Y":
                list_api_data.append(api_dict)
        return list_api_data
    else:
        log.info("表格是空数据!")
        return None


testdata = read_data()


@myddt.ddt
class CheckResult(unittest.TestCase):
    def compare_code_result(self):
        """1.对比预期code和接口响应返回的status code"""
        for row in range(2, all_rows + 1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            """预期status code和接口返回status code"""
            ex_status_code = case_table_sheet.cell(row=row, column=7).value
            ac_status_code = case_table_sheet.cell(row=row, column=8).value
            """判断两个status code是否相等"""
            if is_run == 'Y' or is_run == 'y':
                if ex_status_code and ac_status_code != '':
                    if ex_status_code == ac_status_code:
                        case_table_sheet.cell(row=row, column=9, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])
                    else:
                        case_table_sheet.cell(row=row, column=9, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])  # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=13, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])  # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                        (case_table_sheet.cell(row=row, column=2).value,
                                                                         case_table_sheet.cell(row=row, column=7).value,
                                                                         case_table_sheet.cell(row=row,
                                                                                               column=8).value))
                else:
                    log.error("第 %d 行 status_code为空" % row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！" % row)
        case_table.save(cases_dir)

    def compare_text_result(self):
        """对比预期response和实际返回的response.text，根据预期和实际结果的关系进行处理"""
        for row in range(2, all_rows + 1):
            """接口返回的response.text"""
            response_text = case_table_sheet.cell(row=row, column=12).value
            response_text_dict = dict_res(response_text)
            """预期结果"""
            expect_text = case_table_sheet.cell(row=row, column=10).value
            """接口关键字"""
            key_word = case_table_sheet.cell(row=row, column=3).value
            """status_code对比结果"""
            code_result = case_table_sheet.cell(row=row, column=9).value
            """预期text和response.text的关系"""
            relation = case_table_sheet.cell(row=row, column=11).value
            is_run = case_table_sheet.cell(row=row, column=16).value
            if is_run == 'Y' or is_run == 'y':
                """
                1.status_code 对比结果pass的前提下，判断response.text断言是否正确,
                2.status_code 对比结果fail时，用例整体结果设为fail
                """
                if code_result == 'pass':
                    if key_word in ('create', 'query', 'update', 'delete'):
                        self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                    else:
                        log.error("请确认第%d行的key_word" % row)
                elif code_result == 'fail':
                    case_table_sheet.cell(row=row, column=14, value='fail').fill = PatternFill('solid', fgColor=
                    colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value,
                                                                     case_table_sheet.cell(row=row, column=7).value,
                                                                     case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("请确认第 %d 行 status_code对比结果" % row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！" % row)

        case_table.save(cases_dir)

    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        """
        :param key_word:
        :param relation:
        :param expect_text:
        :param response_text:
        :param response_text_dict:
        :param row:
        :param column:
        :return: 根据expect_text, response_text的关系，进行断言, 目前只处理了等于和包含两种关系
        """
        if key_word == 'create':
            if relation == '=':
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']),
                                             '第%d行的response_text长度和预期不一致' % row)
                        except:
                            # log.info("第 %d 行 response_text返回的id和预期id长度不一致" %row)
                            case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               3])
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '第%d行的expect_text:%s和response_text:%s不相等' % (
                                row, expect_text, response_text))
                        except:
                            # log.info("第%d行的expect_text:%s和response_text:%s不相等" %(row,expect_text, response_text))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               3])
                else:
                    '''只返回一个id串的情况下，判断预期长度和id长度一致'''
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行的response_text长度和预期不一致' % row)
                    except:
                        # log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])

            elif relation == 'in':
                """返回多内容时，断言多个值可以用&连接，并且expect_text包含在response_text中"""
                if "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' % (row, i))
                        except:
                            # log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               3])
                else:
                    try:
                        self.assertIn(expect_text, response_text,
                                      '第 %d 行 预期结果：%s没有包含在response_text中' % (row, expect_text))
                    except:
                        # log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])
            else:
                # log.info("请确认第 %d 行 预期expect_text和response_text的relatrion" %row)
                case_table_sheet.cell(row=row, column=column, value='请确认%d行 的预期text和接口response.text的relatrion' % row)
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}',
                                            '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行expect_text和response_text不相等' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])
                elif expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                    colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertEqual(expect_text, response_text,
                                         '第%d行expect_text:%s和response_text:%s不相等' % (row, expect_text, response_text))
                    except:
                        # log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])

            elif relation == 'in':
                if "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' % (row, i))
                        except:
                            # log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid',
                                                                                                           fgColor=
                                                                                                           colors.COLOR_INDEX[
                                                                                                               3])
                else:
                    try:
                        self.assertIn(expect_text, response_text,
                                      '第 %d 行 预期结果：%s没有包含在response_text中' % (row, expect_text))
                    except:
                        # log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill = PatternFill('solid', fgColor=
                        colors.COLOR_INDEX[3])
            else:
                # log.info("请确认第 %d 行 预期expect_text和response_text的relation" % row)
                case_table_sheet.cell(row=row, column=column,
                                      value='请确认第 %d 行 预期expect_text和response_text的relation' % row)
        else:
            log.error("请确认第 %d 行 的key_word" % row)
        case_table.save(cases_dir)

    def deal_result(self):
        """
        对比code
        self.compare_code_result()
        对比text
        self.compare_text_result()
        根据code result和text result判断case最终结果
        :return: 对比case最终的结果
        """
        self.compare_code_result()
        self.compare_text_result()
        for row in range(2, all_rows + 1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            status_code_result = case_table_sheet.cell(row=row, column=9).value
            response_text_result = case_table_sheet.cell(row=row, column=13).value
            if is_run == 'Y' or is_run == 'y':
                if status_code_result == 'pass' and response_text_result == 'pass':
                    # log.info("测试用例-%s pass" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='pass').fill = PatternFill('solid', fgColor=
                    colors.COLOR_INDEX[3])
                    case_table_sheet.cell(row=row, column=15, value='')
                elif status_code_result == 'fail':
                    # log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill = PatternFill('solid', fgColor=
                    colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code对比失败,预期为%s,实际为%s' \
                                                                    % (case_table_sheet.cell(row=row, column=2).value,
                                                                       case_table_sheet.cell(row=row, column=7).value,
                                                                       case_table_sheet.cell(row=row, column=8).value))
                elif status_code_result == 'pass' and response_text_result == 'fail':
                    # log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill = PatternFill('solid', fgColor=
                    colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回内容对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value,
                                                                     case_table_sheet.cell(row=row, column=10).value,
                                                                     case_table_sheet.cell(row=row, column=12).value))
                else:
                    log.error("请确认status code或response.text对比结果")
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！" % row)
        case_table.save(cases_dir)

    @myddt.data(*testdata)
    def test_api(self,data):
        self.case_name = data['case_detail']
        self.url=host+data['url']
        self.method=data['method']
        self.case_result = data['case_result']
        self.result2 = data['result2']
        self.header=get_headers()
        self.body=data['parameters']
        self.expect_text = data['expect_text']
        self.extract_data=data['response_text']
        self.readData_code =data["response__status_code"]
        print("******* 执行用例 ->{0} *********".format(self.case_name))
        print("请求URL: {0}".format(self.url))
        print("请求方式: {0}".format(self.method))
        print("请求header:{0}".format(self.header))
        print("请求body:{0}".format(self.body))
        if self.case_result == 'pass':
            print("返回状态码：%s 响应信息：%s" % (self.readData_code,self.extract_data))
            self.assertIn(self.expect_text,self.extract_data,"返回实际结果是->:%s" % self.extract_data)
        else:
            print("返回状态码：%s 响应信息：%s" % (self.readData_code, self.extract_data))
            self.assertIn(self.expect_text, self.extract_data, "返回实际结果是->:%s" % self.extract_data)
