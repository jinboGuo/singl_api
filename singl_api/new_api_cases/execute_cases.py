# coding:utf-8
import json
import os
import random
import re
import time
import unittest
import requests
import xlrd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import  colors,PatternFill
from basic_info.get_auth_token import get_headers, get_headers_root, get_auth_token
from basic_info.ready_dataflow_data import get_dataflow_data, get_executions_data, query_dataflow_data, get_schedulers_data
from basic_info.setting import baymax_sheet, host, baymax_cases_dir, log, resource_type
from httpop.Httpop import Httpop
from new_api_cases.deal_parameters import deal_parameters
from new_api_cases.dw_prepare_datas import sql_analyse_data
from new_api_cases.get_statementId import get_sql_analyse_statement_id, get_sql_analyse_dataset_info, get_sql_execte_statement_id, steps_sql_parseinit_statemenId, \
    steps_sql_analyzeinit_statementId, get_step_output_init_statementId, get_step_output_ensure_statementId, \
    step_sql_analyse_data, step_sql_analyse_flow
from new_api_cases.prepare_datas_for_cases import filesets_data, get_import_dataflow, update_db_driver, \
    update_rtcjob_setting, update_custom_step, get_import_data, get_scheduler_online_data
from util import myddt
from util.comm_util import operateKafka
from util.format_res import dict_res, get_time
from util.get_deal_parameter import get_resourceid

cases_dir = baymax_cases_dir
case_table = load_workbook(cases_dir)
baymax_master=baymax_sheet
case_table_sheet = case_table.get_sheet_by_name(baymax_master)
all_rows = case_table_sheet.max_row
fileset_dir = os.path.join(os.path.abspath('.'),'attachment\\Capture001.png').replace('\\','/')
jar_dir = os.path.join(os.path.abspath('.'),'attachment\\woven-common-1.5.2.jar').replace('\\','/')
jar_driver = os.path.join(os.path.abspath('.'),'attachment\\mysql-connector-java-8.0.2801.jar').replace('\\','/')
jar_custom = os.path.join(os.path.abspath('.'),'attachment\\merce-custom-rtc-steps-1.2.4-Filter.jar').replace('\\','/')
woven_dataflow = os.path.join(os.path.abspath('.'),'attachment\\import_dataflow_steps.woven').replace('\\','/')
woven_dir = os.path.join(os.path.abspath('.'),'attachment\\import_auto_apitest_df.woven').replace('\\','/')
multi_sink_steps = os.path.join(os.path.abspath('.'),'attachment\\mutil_sink_storage.woven').replace('\\','/')
multi_rtc_steps = os.path.join(os.path.abspath('.'),'attachment\\multi_rtc_steps.woven').replace('\\','/')
multi_wf_steps = os.path.join(os.path.abspath('.'),'attachment\\multi_wf_steps.woven').replace('\\','/')
organization = os.path.join(os.path.abspath('.'),'attachment\\import_organization.xlsx').replace('\\','/')
user = os.path.join(os.path.abspath('.'),'attachment\\import_user.xlsx').replace('\\','/')
menu = os.path.join(os.path.abspath('.'),'attachment\\import_menu.xlsx').replace('\\','/')
minio_data = []
httpop = Httpop()
host = host


def deal_request_method():
    """
    判断请求方法，并根据不同的请求方法调用不同的处理方式
    :return:
    """
    try:
        for i in range(2, all_rows+1):
            request_method = case_table_sheet.cell(row=i, column=4).value
            request_method_upper = request_method.upper()
            request_url = host+case_table_sheet.cell(row=i, column=5).value
            old_data = case_table_sheet.cell(row=i, column=6).value
            request_data = deal_parameters(old_data,request_method_upper,request_url)
            log.info("request  data：%s" % request_data)
            api_name = case_table_sheet.cell(row=i, column=1).value
            is_run = case_table_sheet.cell(row=i, column=16).value
            if request_method_upper:
                if is_run =='Y' or is_run=='y':
                    if api_name == 'tenants':
                        """
                        租户的用例需要使用root用户登录后操作
                        根据不同的请求方法，进行分发
                        """
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, host=host, column=8, url=request_url, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, host=host, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers_root())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers_root())
                        else:
                            log.error("请求方法%s不在处理范围内" % request_method)
                    else:
                        """根据不同的请求方法，进行分发"""
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, host=host, column=8, url=request_url, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, host=host, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers())
                        else:
                            log.error("请求方法%s不在处理范围内" % request_method)
                else:
                    log.warn(" 第%d 行脚本未执行，请查看isRun是否为Y或者y！"% i)
            else:
                log.error("第 %d 行请求方法为空" % i)
        '''执行结束后保存表格'''
        case_table.save(cases_dir)
    except Exception as e:
        log.error("执行过程中出错{}".format( e))


def post_request_result_check(row, column, url, host, headers, data, table_sheet_name):
    """
    POST接口请求，脚本里post请求的处理
    :param row:
    :param column:
    :param url:
    :param host:
    :param headers:
    :param data:
    :param table_sheet_name:
    :return:
    """
    global case_detail
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        log.info("请求url：%s" % url)
        if "创建模型woven导入任务" == case_detail:
            files = {"file": ('import_auto_apitest_df.woven',open(woven_dir, 'rb'),"application/octet-stream"),"name":(None,"gjb_type_df_import"+str(random.randint(0, 999999)),None),"remark":(None,"gjb_type_df_import",None),"taskType":(None,"offlineDev",None),"flowResourceId":(None,str(get_resourceid(resource_type[3])),None),"flowResourcePath":(None,"Flows;",None),"datasetResourceId":(None,str(get_resourceid(resource_type[1])),None),"datasetResourcePath":(None,"Datasets;",None),"datasourceResourceId":(None,str(get_resourceid(resource_type[0])),None),"datasourceResourcePath":(None,"Datasources;",None),"schemaResourceId":(None,str(get_resourceid(resource_type[2])),None),"schemaResourcePath":(None,"Schemas;",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型woven文件" == case_detail:
            flow_import_task_info_id,new_data = get_import_data(data)
            new_url = url.format(flow_import_task_info_id)
            log.info("请求url：%s" % new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=get_headers(), data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "创建模型dataflow-woven导入任务" == case_detail:
            files = {"file": ('import_dataflow_steps.woven',open(woven_dataflow, 'rb'),"application/octet-stream"),"name":(None,"gjb_type_scheduler_import"+str(random.randint(0, 999999)),None),"remark":(None,"gjb_type_df_import",None),"taskType":(None,"offlineDev",None),"flowResourceId":(None,str(get_resourceid(resource_type[3])),None),"flowResourcePath":(None,"Flows;",None),"datasetResourceId":(None,str(get_resourceid(resource_type[1])),None),"datasetResourcePath":(None,"Datasets;",None),"datasourceResourceId":(None,str(get_resourceid(resource_type[0])),None),"datasourceResourcePath":(None,"Datasources;",None),"schemaResourceId":(None,str(get_resourceid(resource_type[2])),None),"schemaResourcePath":(None,"Schemas;",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型dataflow-woven文件" == case_detail:
            flow_import_task_info_id,new_data = get_import_data(data)
            new_url = url.format(flow_import_task_info_id)
            log.info("请求url：%s" % new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=get_headers(), data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "创建模型multi-sink_steps导入任务" == case_detail:
            files = {"file": ('mutil_sink_storage.woven',open(multi_sink_steps, 'rb'),"application/octet-stream"),"name":(None,"gjb_type_multi_import"+str(random.randint(0, 999999)),None),"remark":(None,"gjb_type_df_import",None),"taskType":(None,"offlineDev",None),"flowResourceId":(None,str(get_resourceid(resource_type[3])),None),"flowResourcePath":(None,"Flows;",None),"datasetResourceId":(None,str(get_resourceid(resource_type[1])),None),"datasetResourcePath":(None,"Datasets;",None),"datasourceResourceId":(None,str(get_resourceid(resource_type[0])),None),"datasourceResourcePath":(None,"Datasources;",None),"schemaResourceId":(None,str(get_resourceid(resource_type[2])),None),"schemaResourcePath":(None,"Schemas;",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型multi-sink_steps文件" == case_detail:
            flow_import_task_info_id,new_data = get_import_data(data)
            new_url = url.format(flow_import_task_info_id)
            log.info("请求url：%s" % new_url)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=get_headers(), data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "创建模型multi-rtc_steps导入任务" == case_detail:
            files = {"file": ('multi_rtc_steps.woven',open(multi_rtc_steps, 'rb'),"application/octet-stream"),"name":(None,"gjb_type_rtc_import"+str(random.randint(0, 999999)),None),"remark":(None,"gjb_type_df_import",None),"taskType":(None,"offlineDev",None),"flowResourceId":(None,str(get_resourceid(resource_type[3])),None),"flowResourcePath":(None,"Flows;",None),"datasetResourceId":(None,str(get_resourceid(resource_type[1])),None),"datasetResourcePath":(None,"Datasets;",None),"datasourceResourceId":(None,str(get_resourceid(resource_type[0])),None),"datasourceResourcePath":(None,"Datasources;",None),"schemaResourceId":(None,str(get_resourceid(resource_type[2])),None),"schemaResourcePath":(None,"Schemas;",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型multi-rtc_steps文件" == case_detail:
            new_data = get_import_dataflow(headers, host, data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=get_headers(), data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "创建模型multi_wf_steps导入任务" == case_detail:
            files = {"file": ('multi_wf_steps.woven',open(multi_wf_steps, 'rb'),"application/octet-stream"),"name":(None,"gjb_type_wf_import"+str(random.randint(0, 999999)),None),"remark":(None,"gjb_type_wf_import",None),"taskType":(None,"workflow",None),"flowResourceId":(None,str(get_resourceid(resource_type[3])),None),"flowResourcePath":(None,"Flows;",None),"datasetResourceId":(None,str(get_resourceid(resource_type[1])),None),"datasetResourcePath":(None,"Datasets;",None),"datasourceResourceId":(None,str(get_resourceid(resource_type[0])),None),"datasourceResourcePath":(None,"Datasources;",None),"schemaResourceId":(None,str(get_resourceid(resource_type[2])),None),"schemaResourcePath":(None,"Schemas;",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "导入模型multi_wf_steps文件" == case_detail:
            new_data = get_import_dataflow(headers, host, data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=get_headers(), data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif case_detail == '获取SQL执行任务结果':
            exec_use_params = get_sql_analyse_dataset_info(host, data)
            exec_use_params = json.dumps(exec_use_params, separators=(',', ':'))
            statement_id,session_id,cluster_id = get_sql_execte_statement_id(host, data)
            new_url = url.format(statement_id,session_id,cluster_id)
            log.info("request   url：%s " % new_url)
            response = requests.post(url=new_url, headers=headers, data=exec_use_params)
            count_num = 0
            while "waiting" in response.text or "running" in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=new_url, headers=headers, data=exec_use_params)
                time.sleep(5)
                count_num += 1
                if count_num == 50:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '获取step的输出字段分析任务的statementID':
            new_data = step_sql_analyse_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '获取初始化确认step任务的statementID':
            new_data = step_sql_analyse_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '获取Sql语句解析表名,初始化ParseSql任务的statementID' == case_detail:
            new_data = step_sql_analyse_flow(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '获取初始化AnalyzeSq,返回statementId' == case_detail:
            new_data = step_sql_analyse_flow(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建作业woven-dataflow':
            new_data = get_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            time.sleep(5)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上线作业woven-dataflow':
            new_data = get_scheduler_online_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            time.sleep(5)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '查找作业记录executions-woven-dataflow':
            new_data = query_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            time.sleep(5)
            while '"statusType":"READY"' in response.text or '"list":[]' in response.text or '"statusType":"RUNNING"' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=new_data)
                time.sleep(6)
                count_num += 1
                if count_num == 90:
                    break
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建作业woven-rtcflow':
            new_data = get_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上线作业woven-rtcflow':
            new_data = get_scheduler_online_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            time.sleep(5)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '查找作业记录executions-woven-rtcflow':
            new_data = query_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            time.sleep(5)
            while "waiting" in response.text or "READY" in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=new_data)
                time.sleep(5)
                count_num += 1
                if count_num == 70:
                    break
            else:
                log.info("开始往kafka插入数据...")
                operateKafka().send_str_kafka()
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建作业woven-workflow':
            new_data = get_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            time.sleep(5)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上线作业woven-workflow':
            new_data = get_scheduler_online_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            time.sleep(5)
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '查找作业记录executions-woven-workflow':
            new_data = query_dataflow_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            time.sleep(5)
            while '"statusType":"READY"' in response.text or '"list":[]' in response.text or '"statusType":"RUNNING"' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=new_data)
                time.sleep(6)
                count_num += 1
                if count_num == 90:
                    break
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '查询指定流程计划' in case_detail:
            new_data = get_schedulers_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '查询流程执行历史' in case_detail:
            new_data = get_executions_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '根据execution id查询输出结果' == case_detail:
            new_data = get_executions_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '批量删除execution':
            data=json.dumps(data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '初始化Sql'in case_detail:
            new_data = sql_analyse_data(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '指定目录下创建子目录':
            response = requests.post(url=url, headers=headers, json=dict_res(data))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '项目目录指定目录创建dataset':
            response = requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '数据源状态监控分析图数据':
            data = {"fieldList":[{"fieldName":"createTime","fieldValue":get_time(),"comparatorOperator":"GREATER_THAN","logicalOperator":"AND"},{"fieldName":"createTime","fieldValue":1555516800000,"comparatorOperator":"LESS_THAN"}],"sortObject":{"field":"lastModifiedTime","orderDirection":"DESC"},"offset":0,"limit":8,"groupBy":"testTime"}
            response = requests.post(url=url,headers=headers, json=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传驱动文件':
            files = {"file": ('mysql-connector-java-8.0.2801.jar',open(jar_driver, 'rb'),"application/octet-stream"),"dbType":(None,"Mysql",None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传选择器文件':
            files = {"file": ('woven-common-1.5.2.jar',open(jar_dir, 'rb'),"application/octet-stream")}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传自定义节点文件':
            files = {"file": ('merce-custom-rtc-steps-1.2.4-Filter.jar',open(jar_custom, 'rb'),"application/octet-stream")}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '创建非结构化文件集合' in case_detail:
            new_data = filesets_data(data)
            response=requests.post(url=url, headers=headers, json=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "服务器创建新目录" in case_detail:
            new_url=url.format(data)
            minio_data.append(data)
            if "MINIO" in case_detail:
                data={"password":"inforefiner","port":"9000","host":"192.168.1.81","region":"","username":"minio"}
            elif "OZONE" in case_detail:
                data= {}
            response = requests.post(url=new_url, headers=headers, json=data)
            log.info("response data：%s %s" % (response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text)))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "上传图片_MINIO" in case_detail:
            fs = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            response = requests.post(url=url, headers=headers, files=fs)
            log.info("response data：%s %s" % (response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text)))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif "上传图片_OZONE" in case_detail:
            fs = {"file": open(fileset_dir, 'rb')}
            headers.pop('Content-Type')
            response = requests.post(url=url, headers=headers, files=fs)
            log.info("response data：%s %s" % (response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text)))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif case_detail == '导入组织机构模板':
            files = {"file": ('import_organization.xlsx',open(organization, 'rb'),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '导入用户模板':
            files = {"file": ('import_user.xlsx',open(user, 'rb'),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '导入权限模板':
            files = {"file": ('import_menu.xlsx',open(menu, 'rb'),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '获取探索数据集sql过滤配置' == case_detail:
            headers.pop('Content-Type')
            headers["Content-Type"]="application/x-www-form-urlencoded"
            data ={"sqls":"SELECT count(*) FROM `gjb_ttest_hdfs042219`"}
            print(headers, data)
            response=requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '解析探索数据集sql' == case_detail:
            headers.pop('Content-Type')
            headers["Content-Type"]="application/x-www-form-urlencoded"
            data ={"sql": "SELECT count(*) FROM `gjb_ttest_hdfs042219`"}
            print(headers, data)
            response=requests.post(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '监控探索离线任务执行成功' == case_detail:
            #new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("data：%s" % data)
            response = requests.post(url=url, headers=headers, json=dict_res(data))
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            time.sleep(6)
            while '"status":"wait"' in response.text or '"status":"running"' in response.text or '"list":[]' in response.text:
                log.info("再次查询：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, json=dict_res(data))
                time.sleep(6)
                count_num += 1
                if count_num == 60:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            if data:
                if "{}" in url:
                    if "##" in data:
                        new_url = url.format(data.split("##")[1])
                        log.info("请求new_url：%s" % new_url)
                        #print(type(data.split("##")[0]),data.split("##")[0])
                        response = requests.post(url=new_url, headers=headers, data=data.split("##")[0].encode("utf-8"))
                        log.info("response data：%s %s" % (response.status_code, response.text))
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    else:
                        if "/api/sys/meta/explore/sql/executeresult" in url:
                            new_url = url.format(data[1], data[2], data[3],data[0])
                            log.info("请求new_url：%s" % new_url)
                            new_data=[]
                            response = requests.post(url=new_url, headers=headers, data=json.dumps(new_data))
                            count_num = 0
                            while "waiting" in response.text or "running" in response.text:
                                response = requests.post(url=new_url, headers=headers, data=json.dumps(new_data))
                                time.sleep(5)
                                count_num += 1
                                if count_num == 50:
                                    return
                            log.info("response data：%s %s" % (response.status_code, response.text))
                            clean_vaule(table_sheet_name, row, column)
                            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                        else:
                            new_url = url.format(data[1], data[2],data[3])
                            log.info("请求new_url：%s" % new_url)
                            response = requests.post(url=new_url, headers=headers, data=json.dumps(data[0]))
                            count_num = 0
                            while "waiting" in response.text or "running" in response.text:
                                response = requests.post(url=new_url, headers=headers, data=json.dumps(data[0]))
                                time.sleep(5)
                                count_num += 1
                                if count_num == 50:
                                    return
                            log.info("response data：%s %s" % (response.status_code, response.text))
                            clean_vaule(table_sheet_name, row, column)
                            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif str(data).startswith('{') and str(data).endswith('}'):
                    data_dict = dict_res(data)
                    time.sleep(3)
                    response = requests.post(url=url, headers=headers, json=data_dict)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif str(data).startswith('[') and str(data).endswith(']'):
                    response = requests.post(url=url, headers=headers, data=str(data))
                    time.sleep(3)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    log.error("请求体数据错误！")
            else:
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(2)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例:{} 执行过程中出错{}".format(case_detail, e))



def get_request_result_check(url, headers, host, data, table_sheet_name, row, column):
    """
    GET接口请求，GET请求需要从parameter中获取参数,并把参数拼装到URL中
    :param url:
    :param headers:
    :param host:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :return:
    """
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("开始执行：%s" % case_detail)
    log.info("请求url：%s" % url)
    try:
        if data:
            if case_detail == '根据statement id,获取Sql Analyze结果(获取输出字段)':
                statement_id,session_id,cluster_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(statement_id,session_id,cluster_id)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 0
                while "waiting" in response.text or "running" in response.text:
                    log.info("再次查询前：%s %s" % (response.status_code, response.text))
                    response = requests.get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 100:
                        return
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '结束指定statementId对应的查询任务':
                statement_id,session_id,cluster_id = get_sql_analyse_statement_id(host, data)
                new_url = url.format(statement_id,session_id,cluster_id)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '根据解析sql parse接口返回的statementId,获取dataset name':
                datasetName_statementId = steps_sql_parseinit_statemenId(host,data)
                new_url = url.format(datasetName_statementId)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 0
                while "waiting" in response.text or "running" in response.text:
                    response = requests.get(url=new_url, headers=headers)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    time.sleep(5)
                    count_num += 1
                    if count_num == 50:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '根据Sql Analyze返回的statementId,获取SqlAnalyze结果':
                steps_sql_analyse_statementId = steps_sql_analyzeinit_statementId(host,data)
                new_url = url.format(steps_sql_analyse_statementId)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 0
                while "waiting" in response.text or "running" in response.text:
                    response = requests.get(url=new_url, headers=headers)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    time.sleep(5)
                    count_num += 1
                    if count_num == 50:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '导出flow':
                token = get_auth_token()
                new_url = url.format(token)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '查看指定id数据源':
                new_url = url.format(data)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '根据statementID获取step的输出字段':
                init_statement_id = get_step_output_init_statementId(host, data)
                new_url = url.format(init_statement_id)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 1
                while "running" in response.text or "waiting" in response.text:
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    time.sleep(5)
                    response = requests.get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 50:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '根据statementID确认step':
                ensure_statement_id = get_step_output_ensure_statementId(host, data)
                new_url = url.format(ensure_statement_id)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                count_num = 0
                while "running" in response.text or "waiting" in response.text:
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    time.sleep(5)
                    response = requests.get(url=new_url, headers=headers)
                    count_num += 1
                    if count_num == 50:
                        return
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif "运行成功" in case_detail:
                new_url=url.format(data)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url,headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                status=json.loads(response.text)["statusType"]
                while status == "WAITTING" or status =="RUNNING" or status =="READY":
                      #log.info("------进入while循环------\n")
                      response = requests.get(url=new_url,headers=headers)
                      status=json.loads(response.text)["statusType"]
                      log.info("------再次查询后的状态为: %s------\n" % status)
                      time.sleep(10)
                if status == "SUCCEEDED":
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    #log.info("flow执行状态出错")
                    return
            elif 'datasetId不存在'in case_detail:
                response = requests.get(url=url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '查询详情'in case_detail:
                new_url = url.format(data)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                new_url = url.format(data)
                log.info('new_url:%s' % new_url)
                response = requests.get(url=new_url, headers=headers)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            response = requests.get(url=url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例：{} 执行过程中出错{}".format(case_detail, e))



def put_request_result_check(url, row, data, table_sheet_name, column, headers):
    """
    PUT接口请求
    :param url:
    :param row:
    :param data:
    :param table_sheet_name:
    :param column:
    :param headers:
    :return:
    """
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("开始执行：%s" % case_detail)
    log.info("请求url：%s" % url)
    try:
        if case_detail == '项目目录改名':
            response = requests.put(url=url, headers=headers, data=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(table_sheet_name, row, column, response.status_code)
            write_result(table_sheet_name, row, column+4, response.text)
        elif '更新数据库驱动' == case_detail:
            new_data = update_db_driver(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '更新自定义节点':
            new_data, step_id = update_custom_step(data)
            new_url = url.format(step_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '更新rtcflow参数模板':
            new_data = update_rtcjob_setting(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.put(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '&&' in str(data):
            '''分隔参数'''
            parameters = data.split('&&')
            '''拼接URL'''
            new_url = url.format(parameters[0])
            log.info("new_url：%s" % new_url)
            '''发送的参数体'''
            parameters_data = parameters[1]
            if parameters_data.startswith('{'):
                response = requests.put(url=new_url, headers=headers, json=dict_res(parameters_data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column+4, response.text)
            else:
                log.error('请确认第%d行parameters中需要update的值格式，应为id&{data}' % row)
        else:
            if "{}" in url:
                new_url = url.format(data["id"])
                log.info("new_url：%s" % new_url)
                response = requests.put(url=new_url, headers=headers,data=json.dumps(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
            else:
                response = requests.put(url=url, headers=headers, json=dict_res(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
    except Exception as e:
        log.error("测试用例：{} 执行过程中出错{}".format(case_detail,e))


def delete_request_result_check(url, data, table_sheet_name, row, column, headers):
    """
    delete接口请求
    :param url:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :param headers:
    :return:
    """
    case_detail = case_table_sheet.cell(row=row, column=2).value
    log.info("开始执行：%s" % case_detail)
    log.info("请求url：%s" % url)
    try:
        if case_detail == '删除非结构化文件集合':
            response = requests.delete(url=url, headers=headers,json=data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif "服务器删除新目录" in case_detail:
            if "_MINIO" in case_detail:
                data={"conf":{"password":"inforefiner","port":"9000","host":"192.168.1.81","region":"","username":"minio"},"name":minio_data}
            elif "_OZONE" in case_detail:
                data={"conf":{},"name":minio_data}
            response = requests.delete(url=url, headers=headers, json=data)
            log.info("response data：%s %s" % (response.status_code, ILLEGAL_CHARACTERS_RE.sub(r'', response.text)))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=ILLEGAL_CHARACTERS_RE.sub(r'', response.text))
        elif isinstance(data, str):
            log.info("data：%s" % data)
            if "{}" in url:
                if "##" in data:
                    new_url = url.format(data.split("##")[0])
                    log.info("new_url：%s" % new_url)
                    response = requests.delete(url=new_url, headers=headers ,data=data.split("##")[1])
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    new_url = url.format(data)
                    log.info("new_url：%s" % new_url)
                    response = requests.delete(url=new_url, headers=headers)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                response = requests.delete(url=url, headers=headers,data=json.dumps(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(table_sheet_name, row, column, response.status_code)
                write_result(table_sheet_name, row, column + 4, response.text)
        elif isinstance(data,list):
            response = requests.delete(url=url, headers=headers,data=json.dumps(data))
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            response = requests.delete(url=url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例：{} 执行过程中出错{}".format(case_detail,e))



def write_result(sheet, row, column, value):
    """
    写入返回结果
    :param sheet:
    :param row:
    :param column:
    :param value:
    :return: 写入返回结果
    """
    sheet.cell(row=row, column=column, value=value)



def clean_vaule(sheet, row, column):
    """
    :param sheet:
    :param row:
    :param column:
    :return: 写入结果前，先把结果和对比结果全部清空
    """
    sheet.cell(row=row, column=column, value='')
    sheet.cell(row=row, column=column+1, value='')
    sheet.cell(row=row, column=column + 4, value='')
    sheet.cell(row=row, column=column + 5, value='')
    sheet.cell(row=row, column=column + 6, value='')
    sheet.cell(row=row, column=column + 7, value='')


def read_data():
        data = xlrd.open_workbook(cases_dir)
        table = data.sheet_by_name(baymax_master)
        """获取总行数"""
        n_rows = table.nrows
        if n_rows > 1:
            """获取第一行的内容，列表格式"""
            keys = table.row_values(0)
            list_api_data = []
            """获取每一行的内容，列表格式"""
            for col in range(1, n_rows):
                values = table.row_values(col)
                """ keys，values组合转换为字典"""
                api_dict = dict(zip(keys, values))
                if api_dict['is_run']=="y" or api_dict['is_run']=="Y":
                    list_api_data.append(api_dict)
            return list_api_data
        else:
            #log.info("表格是空数据!")
            return None

testdata = read_data()


@myddt.ddt
class CheckResult(unittest.TestCase):
    def compare_code_result(self):
        """1.对比预期code和接口响应返回的status code"""
        for row in range(2, all_rows+1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            """预期status code和接口返回status code"""
            ex_status_code = case_table_sheet.cell(row=row, column=7).value
            ac_status_code = case_table_sheet.cell(row=row, column=8).value
            """判断两个status code是否相等"""
            if is_run == 'Y' or is_run == 'y':
                if ex_status_code and ac_status_code != '':
                    if ex_status_code == ac_status_code:
                        case_table_sheet.cell(row=row, column=9, value='pass').fill = PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        case_table_sheet.cell(row=row, column=9, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=13, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                        (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("第 %d 行 status_code为空" %row)
            else:
                log.warn("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    def compare_text_result(self):
        """对比预期response和实际返回的response.text，根据预期和实际结果的关系进行处理"""
        for row in range(2, all_rows+1):
            """接口返回的response.text"""
            response_text = case_table_sheet.cell(row=row, column=12).value
            response_text_dict = dict_res(response_text)
            """预期结果"""
            expect_text = case_table_sheet.cell(row=row, column=10).value
            """接口关键字"""
            key_word = case_table_sheet.cell(row=row, column=3).value
            """status_code对比结果"""
            code_result = case_table_sheet.cell(row=row, column=9).value
            """预期text和response.text的关系"""
            relation = case_table_sheet.cell(row=row, column=11).value
            is_run = case_table_sheet.cell(row=row, column=16).value
            if is_run == 'Y' or is_run == 'y':
                """
                1.status_code 对比结果pass的前提下，判断response.text断言是否正确,
                2.status_code 对比结果fail时，用例整体结果设为fail
                """
                if code_result == 'pass':
                    if key_word in ('create', 'query', 'update', 'delete'):
                        self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                    else:
                        log.error("请确认第%d行的key_word" % row)
                elif code_result == 'fail':
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("请确认第 %d 行 status_code对比结果" % row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)

        case_table.save(cases_dir)


    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        """
        :param key_word:
        :param relation:
        :param expect_text:
        :param response_text:
        :param response_text_dict:
        :param row:
        :param column:
        :return: 根据expect_text, response_text的关系，进行断言, 目前只处理了等于和包含两种关系
        """
        if key_word == 'create':
            if relation == '=':
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']), '第%d行的response_text长度和预期不一致' % row)
                        except:
                            #log.info("第 %d 行 response_text返回的id和预期id长度不一致" %row)
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '第%d行的expect_text:%s和response_text:%s不相等' % (row,expect_text, response_text))
                        except:
                            #log.info("第%d行的expect_text:%s和response_text:%s不相等" %(row,expect_text, response_text))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    '''只返回一个id串的情况下，判断预期长度和id长度一致'''
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行的response_text长度和预期不一致' % row)
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                """返回多内容时，断言多个值可以用&连接，并且expect_text包含在response_text中"""
                if expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relatrion" %row)
                case_table_sheet.cell(row=row, column=column, value='请确认%d行 的预期text和接口response.text的relatrion'%row)
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}', '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行expect_text和response_text不相等' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertEqual(expect_text, response_text, '第%d行expect_text:%s和response_text:%s不相等' % (row,expect_text,response_text))
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                if expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relation" % row)
                case_table_sheet.cell(row=row, column=column, value='请确认第 %d 行 预期expect_text和response_text的relation'%row)
        else:
            log.error("请确认第 %d 行 的key_word" % row)
        case_table.save(cases_dir)


    def deal_result(self):
        """
        对比code
        self.compare_code_result()
        对比text
        self.compare_text_result()
        根据code result和text result判断case最终结果
        :return: 对比case最终的结果
        """
        self.compare_code_result()
        self.compare_text_result()
        for row in range(2, all_rows + 1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            status_code_result = case_table_sheet.cell(row=row, column=9).value
            response_text_result = case_table_sheet.cell(row=row, column=13).value
            if is_run=='Y' or is_run=='y':
                if status_code_result == 'pass' and response_text_result == 'pass':
                    #log.info("测试用例-%s pass" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    case_table_sheet.cell(row=row, column=15, value='')
                elif status_code_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code对比失败,预期为%s,实际为%s' \
                                                                    % (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                elif status_code_result == 'pass' and response_text_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回内容对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=10).value, case_table_sheet.cell(row=row, column=12).value))
                else:
                    log.error("请确认status code或response.text对比结果")
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    @myddt.data(*testdata)
    def test_api(self,data):
        self.case_name = data['case_detail']
        self.url=host+data['url']
        self.method=data['method']
        self.case_result = data['case_result']
        self.result2 = data['result2']
        self.header=get_headers()
        self.body=data['parameters']
        self.expect_text = data['expect_text']
        self.extract_data=data['response_text']
        self.readData_code =data["response__status_code"]
        print("******* 执行用例 ->{0} *********".format(self.case_name))
        print("请求URL: {0}".format(self.url))
        print("请求方式: {0}".format(self.method))
        print("请求header:{0}".format(self.header))
        print("请求body:{0}".format(self.body))
        if self.case_result == 'pass':
            print("返回状态码：%s 响应信息：%s" % (self.readData_code,self.extract_data))
            self.assertIn(self.expect_text,self.extract_data,"返回实际结果是->:%s" % self.extract_data)
        else:
            print("返回状态码：%s 响应信息：%s" % (self.readData_code, self.extract_data))
            self.assertIn(self.expect_text, self.extract_data, "返回实际结果是->:%s" % self.extract_data)
