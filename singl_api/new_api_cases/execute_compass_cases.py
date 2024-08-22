# coding:utf-8
import json
import os
import re
import time
from openpyxl.styles import PatternFill, colors
from util import myddt
import xlrd
from openpyxl import load_workbook
import requests
from util.format_res import dict_res
from basic_info.setting import compass_host, compass_sheet, compass_cases_dir, log
from basic_info.get_auth_token import get_headers_root, get_headers
from new_api_cases.compass_deal_parameters import deal_parameters
import unittest
from new_api_cases.compass_prepare_datas import update_job, add_job, add_jobSingle, \
    add_jobMap, update_jobMap, move_asset_directory, duplicate_asset_directory, duplicate_move_asset_directory, \
    update_asset_directory, publish_flow, get_jobview_history, execution_task, \
    qa_rule_task, publish_qa_flow, get_qa_jobview_history, delete_asset_directory, query_re_th, \
    query_re_th_ext, get_job_view, get_input_dataset, get_output_dataset, save_node_prop, \
    update_draft, update_info_list, update_setting, get_job_view_exec, add_re_th_ext, update_re_th_ext, update_re_th, \
    dc_collect_group, add_re_th, publish_dsp_flow, get_dsp_jobview_history, update_job_single

cases_dir = compass_cases_dir
case_table = load_workbook(cases_dir)
compass_master=compass_sheet
case_table_sheet = case_table.get_sheet_by_name(compass_master)
all_rows = case_table_sheet.max_row
jar_dir_push = os.path.join(os.path.abspath('.'),'attachment\\Scheduler_import.xlsx').replace('\\','/')
jar_driver = os.path.join(os.path.abspath('.'),'attachment\\mysql-connector-java-8.0.2801.jar').replace('\\','/')
host = compass_host


def deal_request_method():
    """
    判断请求方法，并根据不同的请求方法调用不同的处理方式
    :return:
    """
    try:
        for i in range(2, all_rows+1):
            request_method = case_table_sheet.cell(row=i, column=4).value
            request_method_upper = request_method.upper()
            request_url = host+case_table_sheet.cell(row=i, column=5).value
            old_data = case_table_sheet.cell(row=i, column=6).value
            request_data = deal_parameters(old_data,request_method_upper,request_url)
            log.info("request  data：%s" % request_data)
            api_name = case_table_sheet.cell(row=i, column=1).value
            is_run = case_table_sheet.cell(row=i, column=16).value
            if request_method_upper:
                if is_run =='Y' or is_run=='y':
                    if api_name == 'tenants':
                        """
                        租户的用例需要使用root用户登录后操作
                        根据不同的请求方法，进行分发
                        """
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers_root(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers_root())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers_root())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                    else:
                        """根据不同的请求方法，进行分发"""
                        if request_method_upper == 'POST':
                            post_request_result_check(row=i, column=8, url=request_url, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet)
                        elif request_method_upper == 'GET':
                            get_request_result_check(url=request_url, headers=get_headers(), data=request_data, table_sheet_name=case_table_sheet, row=i, column=8)
                        elif request_method_upper == 'PUT':
                            put_request_result_check(url=request_url, row=i, data=request_data, table_sheet_name=case_table_sheet, column=8, headers=get_headers())
                        elif request_method_upper == 'DELETE':
                            delete_request_result_check(url=request_url, data=request_data, table_sheet_name=case_table_sheet, row=i, column=8, headers=get_headers())
                        else:
                            log.info("请求方法%s不在处理范围内" % request_method)
                else:
                    log.info(" 第%d 行脚本未执行，请查看isRun是否为Y或者y！"% i)
            else:
                log.info("第 %d 行请求方法为空" % i)
        '''执行结束后保存表格'''
        case_table.save(cases_dir)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(i, e))

def post_request_result_check(row, column, url, headers, data, table_sheet_name):
    """
    POST接口请求，脚本里post请求的处理
    :param row:
    :param column:
    :param url:
    :param headers:
    :param data:
    :param table_sheet_name:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        log.info("请求url：%s" % url)
        if '新增调度任务' in case_detail:
            new_data = add_job(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '新增任务类型' in case_detail:
            new_data = add_jobSingle(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '新增数据流程依赖关系' == case_detail:
            new_data = add_jobMap(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '查询资源参数名称' == case_detail:
            new_data = query_re_th(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '新增数据资源参数' == case_detail:
            new_data = add_re_th(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '查询资源参数附加值名称' == case_detail:
            new_data = query_re_th_ext(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '新增数据资源参数附加值' == case_detail:
            new_data = add_re_th_ext(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '导入调度任务':
            files = {"file": open(jar_dir_push, 'rb')}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '资产目录移动':
            parent_id, new_data = move_asset_directory(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            new_url = url.format(parent_id)
            log.info("new_url：%s " % new_url)
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '添加重名资产目录':
            new_data = duplicate_asset_directory(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '新建视图任务':
            new_data = get_job_view(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '上传驱动文件':
            files = {"file": ('mysql-connector-java-8.0.2801.jar', open(jar_driver, 'rb'), "application/octet-stream"),
                     "dbType": (None, "Mysql", None)}
            headers.pop('Content-Type')
            response = requests.post(url=url, files=files, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建采集组':
            new_data = dc_collect_group(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '生成input节点数据集':
            draft_id, new_data = get_input_dataset(data)
            new_url = url.format(draft_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '生成output节点数据集':
            draft_id, new_data = get_output_dataset(data)
            new_url = url.format(draft_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '保存节点属性配置信息':
            draft_id, new_data = save_node_prop(data)
            new_url = url.format(draft_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '更新画布基础信息':
            draft_id, new_data = update_draft(data)
            new_url = url.format(draft_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '更新output节点信息':
            draft_id, new_data = update_info_list(data)
            new_url = url.format(draft_id)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=new_url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail=='监控订阅资源服务执行状态成功':
            log.info("new_data：%s" % data)
            response = requests.post(url=url, headers=headers, data=data)
            count_num = 0
            time.sleep(6)
            while '"isRunning":0' in response.text or '"list":[]' in response.text or '"isRunning":1' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(6)
                count_num += 1
                if count_num == 60:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '创建质量分析任务':
            new_data = qa_rule_task(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail=='监控质检任务执行状态成功':
            log.info("new_data：%s" % data)
            response = requests.post(url=url, headers=headers, data=data)
            count_num = 0
            time.sleep(6)
            while '"statusType":"READY"' in response.text or '"list":[]' in response.text or '"statusType":"RUNNING"' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(6)
                count_num += 1
                if count_num == 60:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '新建视图任务版本':
            new_data = get_jobview_history(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '新建视图任务版本-质量核查':
            new_data = get_qa_jobview_history(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '新建视图任务版本-数据共享':
            new_data = get_dsp_jobview_history(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif case_detail == '监控采集任务执行状态运行成功':
            new_data = execution_task(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            while '"runningStatus":"RUNNING"' in response.text or '"runningStatus":"CREATED"' in response.text or '"list":[]' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=new_data)
                time.sleep(6)
                count_num += 1
                if count_num == 50:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        elif '监控调度任务执行状态成功' in case_detail:
            new_data = get_job_view_exec(data)
            new_data = json.dumps(new_data, separators=(',', ':'))
            log.info("new_data：%s" % new_data)
            response = requests.post(url=url, headers=headers, data=new_data)
            log.info("response data：%s %s" % (response.status_code, response.text))
            count_num = 0
            time.sleep(6)
            while '"status":1' in response.text or '"list":[]' in response.text:
                log.info("再次查询前：%s %s" % (response.status_code, response.text))
                response = requests.post(url=url, headers=headers, data=new_data)
                time.sleep(6)
                count_num += 1
                if count_num == 60:
                    return
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            if data:
                if "{}" in url:
                    new_url = url.format(data[1], data[2], data[3])
                    log.info("请求new_url：%s" % new_url)
                    response = requests.post(url=new_url, headers=headers, data=json.dumps(data[0]))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    count_num = 0
                    while "waiting" in response.text or "running" in response.text:
                        response = requests.post(url=new_url, headers=headers, data=json.dumps(data[0]))
                        time.sleep(5)
                        count_num += 1
                        if count_num == 50:
                            return
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif str(data).startswith('{') and str(data).endswith('}'):
                    data_dict = dict_res(data)
                    response = requests.post(url=url, headers=headers, json=data_dict)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                elif str(data).startswith('[') and str(data).endswith(']'):
                    response = requests.post(url=url, headers=headers, data=str(data))
                    time.sleep(3)
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                    write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                else:
                    log.error("请求体数据错误！")
            else:
                response = requests.post(url=url, headers=headers, data=data)
                time.sleep(1)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("测试用例:{} 执行过程中出错{}".format(case_detail, e))



def get_request_result_check(url, headers, data, table_sheet_name, row, column):
    """
    GET请求需要从parameter中获取参数,并把参数拼装到URL中
    :param url:
    :param headers:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        log.info("请求url：%s" % url)
        if data:
            new_url = url.format(data)
            log.info('new_url:%s' % new_url)
            response = requests.get(url=new_url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
        else:
            response = requests.get(url=url, headers=headers)
            log.info("response data：%s %s" % (response.status_code, response.text))
            clean_vaule(table_sheet_name, row, column)
            write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
            write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


def put_request_result_check(url, row, data, table_sheet_name, column, headers):
    """
    PUT接口请求
                if '&' in str(data):
                # 分隔参数
                parameters = data.split('&')
                # 拼接URL
                new_url = url.format(parameters[0])
    :param url:
    :param row:
    :param data:
    :param table_sheet_name:
    :param column:
    :param headers:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        log.info("请求url：%s" % url)
        if data:
            if case_detail == '编辑调度任务':
                new_data = update_job(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '编辑立即执行任务':
                new_data = update_job_single(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '编辑数据流程依赖关系':
                new_data = update_jobMap(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '编辑资源参数名称':
                re_th_oid, new_data = update_re_th(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(re_th_oid)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '编辑资源参数附加值名称':
                re_th_ext_oid, new_data = update_re_th_ext(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(re_th_ext_oid)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '资产目录重命名':
                new_data = update_asset_directory(data)
                log.info("request   url：%s " % url)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '资产目录重命名-名称重复':
                new_data = duplicate_move_asset_directory(data)
                log.info("request   url：%s " % url)
                new_data = json.dumps(new_data, separators=(',', ':'))
                response = requests.put(url=url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '更新任务配置信息':
                draft_id, new_data = update_setting(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(draft_id)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '发布调度flow':
                job_id, new_data = publish_flow(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(job_id)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '发布调度质量flow':
                job_id, new_data = publish_qa_flow(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(job_id)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif case_detail == '发布调度共享flow':
                job_id, new_data = publish_dsp_flow(data)
                new_data = json.dumps(new_data, separators=(',', ':'))
                new_url = url.format(job_id)
                response = requests.put(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif '&&' in str(data):
                '''分隔参数'''
                parameters = data.split('&&')
                '''拼接URL'''
                new_url = url.format(parameters[0])
                log.info("new_url：%s" % new_url)
                '''发送的参数体'''
                parameters_data = parameters[1]
                if parameters_data.startswith('{'):
                    response = requests.put(url=new_url, headers=headers, json=dict_res(parameters_data))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(table_sheet_name, row, column, response.status_code)
                    write_result(table_sheet_name, row, column + 4, response.text)
                else:
                    log.info('请确认第%d行parameters中需要update的值格式，应为id&{data}' % row)
            else:
                if "{}" in url:
                    new_url = url.format(data["id"])
                    log.info("new_url：%s" % new_url)
                    response = requests.put(url=new_url, headers=headers, data=json.dumps(data))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(table_sheet_name, row, column, response.status_code)
                    write_result(table_sheet_name, row, column + 4, response.text)
                else:
                    response = requests.put(url=url, headers=headers, json=dict_res(data))
                    log.info("response data：%s %s" % (response.status_code, response.text))
                    clean_vaule(table_sheet_name, row, column)
                    write_result(table_sheet_name, row, column, response.status_code)
                    write_result(table_sheet_name, row, column + 4, response.text)
        else:
            log.error("请求体为空！")
    except Exception as e:
        log.error("{}执行过程中出错{}".format(case_detail, e))


def delete_request_result_check(url, data, table_sheet_name, row, column, headers):
    """
    delete接口请求
    :param url:
    :param data:
    :param table_sheet_name:
    :param row:
    :param column:
    :param headers:
    :return:
    """
    try:
        case_detail = case_table_sheet.cell(row=row, column=2).value
        log.info("开始执行：%s" % case_detail)
        log.info("请求url：%s" % url)
        if data:
            if  case_detail=="删除调度目录":
                asset_id, new_data = delete_asset_directory(data)
                log.info("request   data：%s " % new_data)
                new_url = url.format(asset_id)
                log.info("new_url：%s " % new_url)
                response = requests.delete(url=new_url, headers=headers, data=new_data)
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            elif isinstance(data, str):
                    log.info("data：%s" % data)
                    if "{}" in url:
                        new_url = url.format(data)
                        log.info("new_url：%s" % new_url)
                        response = requests.delete(url=new_url, headers=headers)
                        log.info("response data：%s %s" % (response.status_code, response.text))
                        clean_vaule(table_sheet_name, row, column)
                        write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                        write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
                    else:
                        response = requests.delete(url=url, headers=headers, data=json.dumps(data))
                        log.info("response data：%s %s" % (response.status_code, response.text))
                        clean_vaule(table_sheet_name, row, column)
                        write_result(table_sheet_name, row, column, response.status_code)
                        write_result(table_sheet_name, row, column + 4, response.text)
            elif isinstance(data, list):
                response = requests.delete(url=url, headers=headers, data=json.dumps(data))
                log.info("response data：%s %s" % (response.status_code, response.text))
                clean_vaule(table_sheet_name, row, column)
                write_result(sheet=table_sheet_name, row=row, column=column, value=response.status_code)
                write_result(sheet=table_sheet_name, row=row, column=column + 4, value=response.text)
            else:
                log.error("{}执行过程中出错".format(case_detail))
        else:
            log.error("请求体数据为空！")
    except Exception as e:
            log.error("{}执行过程中出错{}".format(case_detail, e))



def write_result(sheet, row, column, value):
    """
    写入返回结果
    :param sheet:
    :param row:
    :param column:
    :param value:
    :return:
    """
    sheet.cell(row=row, column=column, value=value)



def clean_vaule(sheet, row, column):
    """
    写入结果前，先把结果和对比结果全部清空
    :param sheet:
    :param row:
    :param column:
    :return:
    """
    sheet.cell(row=row, column=column, value='')
    sheet.cell(row=row, column=column+1, value='')
    sheet.cell(row=row, column=column + 4, value='')
    sheet.cell(row=row, column=column + 5, value='')
    sheet.cell(row=row, column=column + 6, value='')
    sheet.cell(row=row, column=column + 7, value='')



def read_data():
        data = xlrd.open_workbook(cases_dir)
        table = data.sheet_by_name(compass_master)
        """获取总行数"""
        nrows = table.nrows
        if nrows > 1:
            """获取第一行的内容，列表格式"""
            keys = table.row_values(0)
            list_api_data = []
            """获取每一行的内容，列表格式"""
            for col in range(1, nrows):
                values = table.row_values(col)
                """ keys，values组合转换为字典"""
                api_dict = dict(zip(keys, values))
                if api_dict['is_run']=="y" or api_dict['is_run']=="Y":
                    list_api_data.append(api_dict)
            return list_api_data
        else:
            log.info("表格是空数据!")
            return None

testdata = read_data()



@myddt.ddt
class CheckResult(unittest.TestCase):
    def compare_code_result(self):
        """1.对比预期code和接口响应返回的status code"""
        for row in range(2, all_rows+1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            """预期status code和接口返回status code"""
            ex_status_code = case_table_sheet.cell(row=row, column=7).value
            ac_status_code = case_table_sheet.cell(row=row, column=8).value
            """判断两个status code是否相等"""
            if is_run == 'Y' or is_run == 'y':
                if ex_status_code and ac_status_code != '':
                    if ex_status_code == ac_status_code:
                        case_table_sheet.cell(row=row, column=9, value='pass').fill = PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        case_table_sheet.cell(row=row, column=9, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=13, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2]) # code不等时，用例结果直接判断为失败
                        case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                        (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("第 %d 行 status_code为空" %row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    def compare_text_result(self):
        """对比预期response和实际返回的response.text，根据预期和实际结果的关系进行处理"""
        for row in range(2, all_rows+1):
            """接口返回的response.text"""
            response_text = case_table_sheet.cell(row=row, column=12).value
            response_text_dict = dict_res(response_text)
            """预期结果"""
            expect_text = case_table_sheet.cell(row=row, column=10).value
            """接口关键字"""
            key_word = case_table_sheet.cell(row=row, column=3).value
            """status_code对比结果"""
            code_result = case_table_sheet.cell(row=row, column=9).value
            """预期text和response.text的关系"""
            relation = case_table_sheet.cell(row=row, column=11).value
            is_run = case_table_sheet.cell(row=row, column=16).value
            if is_run == 'Y' or is_run == 'y':
                """
                1.status_code 对比结果pass的前提下，判断response.text断言是否正确,
                2.status_code 对比结果fail时，用例整体结果设为fail
                """
                if code_result == 'pass':
                    if key_word in ('create', 'query', 'update', 'delete'):
                        self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                    else:
                        log.error("请确认第%d行的key_word" % row)
                elif code_result == 'fail':
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回status_code对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                else:
                    log.error("请确认第 %d 行 status_code对比结果" % row)
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)

        case_table.save(cases_dir)


    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        """
        :param key_word:
        :param relation:
        :param expect_text:
        :param response_text:
        :param response_text_dict:
        :param row:
        :param column:
        :return: 根据expect_text, response_text的关系，进行断言, 目前只处理了等于和包含两种关系
        """
        if key_word == 'create':
            if relation == '=':
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']), '第%d行的response_text长度和预期不一致' % row)
                        except:
                            #log.info("第 %d 行 response_text返回的id和预期id长度不一致" %row)
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '第%d行的expect_text:%s和response_text:%s不相等' % (row,expect_text, response_text))
                        except:
                            #log.info("第%d行的expect_text:%s和response_text:%s不相等" %(row,expect_text, response_text))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    '''只返回一个id串的情况下，判断预期长度和id长度一致'''
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行的response_text长度和预期不一致' % row)
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                """返回多内容时，断言多个值可以用&连接，并且expect_text包含在response_text中"""
                if "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relatrion" %row)
                case_table_sheet.cell(row=row, column=column, value='请确认%d行 的预期text和接口response.text的relatrion'%row)
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}', '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行expect_text和response_text不相等' % row)
                    except:
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                elif expect_text == None and response_text == "":
                    case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertEqual(expect_text, response_text, '第%d行expect_text:%s和response_text:%s不相等' % (row,expect_text,response_text))
                    except:
                        #log.info("第 %d 行 response_text和预期text不相等" %row)
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])

            elif relation == 'in':
                if "&&" in expect_text:
                    for i in expect_text.split("&&"):
                        try:
                            self.assertIn(i, response_text, '第 %d 行 预期结果：%s没有包含在response_text中' %(row,i))
                        except:
                            #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,i))
                            case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                            break
                        else:
                            case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                else:
                    try:
                        self.assertIn(expect_text, response_text, '第 %d 行 预期结果：%s没有包含在response_text中'%(row,expect_text))
                    except:
                        #log.info("第 %d 行 预期结果：%s没有包含在response_text中， 结果对比失败" %(row,expect_text))
                        case_table_sheet.cell(row=row, column=column, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    else:
                        case_table_sheet.cell(row=row, column=column, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
            else:
                #log.info("请确认第 %d 行 预期expect_text和response_text的relation" % row)
                case_table_sheet.cell(row=row, column=column, value='请确认第 %d 行 预期expect_text和response_text的relation'%row)
        else:
            log.error("请确认第 %d 行 的key_word" % row)
        case_table.save(cases_dir)


    def deal_result(self):
        """
        对比code
        self.compare_code_result()
        对比text
        self.compare_text_result()
        根据code result和text result判断case最终结果
        :return: 对比case最终的结果
        """
        self.compare_code_result()
        self.compare_text_result()
        for row in range(2, all_rows + 1):
            is_run = case_table_sheet.cell(row=row, column=16).value
            status_code_result = case_table_sheet.cell(row=row, column=9).value
            response_text_result = case_table_sheet.cell(row=row, column=13).value
            if is_run=='Y' or is_run=='y':
                if status_code_result == 'pass' and response_text_result == 'pass':
                    #log.info("测试用例-%s pass" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='pass').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[3])
                    case_table_sheet.cell(row=row, column=15, value='')
                elif status_code_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code对比失败,预期为%s,实际为%s' \
                                                                    % (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
                elif status_code_result == 'pass' and response_text_result == 'fail':
                    #log.info("测试用例-%s fail" % case_table_sheet.cell(row=row, column=2).value)
                    case_table_sheet.cell(row=row, column=14, value='fail').fill=PatternFill('solid', fgColor=colors.COLOR_INDEX[2])
                    case_table_sheet.cell(row=row, column=15, value='')
                    case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回内容对比失败,预期为%s,实际为%s' %
                                                                    (case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=10).value, case_table_sheet.cell(row=row, column=12).value))
                else:
                    log.error("请确认status code或response.text对比结果")
            else:
                log.error("第 %d 行脚本未执行，请查看isRun是否为Y或者y！"%row)
        case_table.save(cases_dir)


    @myddt.data(*testdata)
    def test_api(self,data):
        self.case_name = data['case_detail']
        self.url=host+data['url']
        self.method=data['method']
        self.case_result = data['case_result']
        self.result2 = data['result2']
        self.header=get_headers()
        self.body=data['parameters']
        self.expect_text = data['expect_text']
        self.extract_data=data['response_text']
        self.readData_code =data["response__status_code"]
        print("******* 执行用例 ->{0} *********".format(self.case_name))
        print("请求URL: {0}".format(self.url))
        print("请求方式: {0}".format(self.method))
        print("请求header:{0}".format(self.header))
        print("请求body:{0}".format(self.body))
        if self.case_result == 'pass':
            print("返回状态码：%s 响应信息：%s" % (self.readData_code,self.extract_data))
            self.assertIn(self.expect_text,self.extract_data,"返回实际结果是->:%s" % self.extract_data)
        else:
            print("返回状态码：%s 响应信息：%s" % (self.readData_code, self.extract_data))
            self.assertIn(self.expect_text, self.extract_data, "返回实际结果是->:%s" % self.extract_data)