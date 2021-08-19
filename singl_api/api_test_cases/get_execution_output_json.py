# coding=gbk
from util.Open_DB import MYSQL
from basic_info.get_auth_token import get_headers
from basic_info.setting import MySQL_CONFIG
from util.format_res import dict_res, get_time
from basic_info.setting import host, tenant_id_83
from basic_info.get_execution_log import GetLog
from new_api_cases.get_statementId import statementId_flow_use, preview_result_flow_use
from util.encrypt import parameter_ungzip
from util.get_tenant import get_tenant
import time, random, requests
from openpyxl import load_workbook
import json
import os

abs_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))

# sink_output_info=[{'flow_id': 'f0f97967-d322-479d-a0fa-02a9c4837085', 'execution_id': 'ccf14e04-9d21-470b-8aa2-f184b9d9585b', 'flow_scheduler_id': '59e0698c-f098-4917-aee2-d5f0e9d03056', 'e_final_status': 'SUCCEEDED', 'dataset_id': 'bdcaf156-c066-4ec5-afa4-6f04e2e7678d'}]

class GetCheckoutDataSet(object):
    """
    �����ύexecution����
    ����ִ�н������������ִ�н����ȡ����Ӧ��sink���ݼ�������
    ���Խ�����бȶ��Ƿ���ȷ
    """

    def __init__(self):
        """��ʼ�����ݿ�����"""
        self.ms = MYSQL(MySQL_CONFIG["HOST"], MySQL_CONFIG["USER"], MySQL_CONFIG["PASSWORD"], MySQL_CONFIG["DB"],MySQL_CONFIG["PORT"])
        self.host = host
        self.table = "flow_dataset_info.xlsx"
        self.table_sheet = '84'

    def get_flow_id(self):
        """
        ��ȡflow id����list��ʽ����
        :return: flow_id_list
        """
        print("------��ȡflow id list------")
        flow_table = load_workbook(abs_dir(self.table))
        info_sheet = flow_table.get_sheet_by_name(self.table_sheet)
        max_row = info_sheet.max_row  # ��ȡ����
        flow_id_list = []
        for i in range(2, max_row+1):
            if info_sheet.cell(row=i, column=2).value and len(info_sheet.cell(row=i, column=2).value) > 10:
                flow_id_list.append(info_sheet.cell(row=i, column=2).value)
        flow_id_list = list(set(flow_id_list))
        return flow_id_list

    def get_flow_id_not_distinct(self):
        """
        ��ȡflow id����list��ʽ����
        :return: flow_id_list
        """
        print("------��ȡflow id list------")
        flow_table = load_workbook(abs_dir(self.table))
        info_sheet = flow_table.get_sheet_by_name(self.table_sheet)
        max_row = info_sheet.max_row  # ��ȡ����
        flow_id_list = []
        for i in range(2, max_row+1):
            if info_sheet.cell(row=i, column=2).value and len(info_sheet.cell(row=i, column=2).value) > 10:
                flow_id_list.append(info_sheet.cell(row=i, column=2).value)
        return flow_id_list


    def data_for_create_scheduler(self):
        """
        1. ����flow_id ����flow_name����Ϣ
        2. ���ݲ�ѯ����flow��Ϣ��ƴװ����scheduler����Ҫʹ�õ�data
        :return: data_list
        """
        print("------��װ������������Ҫ��data------\n")
        data_list = []
        flow_id_list = self.get_flow_id()
        for flow_id in flow_id_list:
            try:
                sql = 'select name, flow_type, parameters from merce_flow where id = "%s"' % flow_id
                flow_info = self.ms.ExecuQuery(sql)
                print('flow_info:', flow_info)
            except Exception as e:
                return e
            else:
                try:
                    flow_name = flow_info[0]["name"]
                    flow_type = flow_info[0]["flow_type"]
                    flow_parameters = flow_info[0]["parameters"]
                    arguments_list = []
                    arguments = {}
                    if flow_parameters:  # flow parameters���ڵ������
                        print('parametersû�н�ѹ��ʱ��', flow_parameters)
                        parameters_use = parameter_ungzip(flow_parameters)  # �����ܺ�Ĳ������н��ܺͽ�ѹ������
                        print('parameters��ѹ����', parameters_use)
                        flow_parameters_list = dict_res(parameters_use)
                        if len(flow_parameters_list) > 0:
                            arguments["name"] = flow_parameters_list[0]["name"]
                            arguments["category"] = flow_parameters_list[0]["category"]
                            arguments["value"] = flow_parameters_list[0]["defaultVal"]
                            arguments["refs"] = flow_parameters_list[0]["refs"]
                            arguments["description"] = flow_parameters_list[0]["description"]
                            arguments_list.append(arguments)
                        # print('arguments:', arguments)
                except KeyError as e:
                    raise e
                except IndexError as T:
                    raise T
            data = {
                "configurations": {
                    "arguments": arguments_list,
                    "properties": [
                        {
                            "name": "all.debug",
                            "value": "false"
                        },
                        {
                            "name": "all.dataset-nullable",
                            "value": "false"
                        },
                        {
                            "name": "all.lineage.enable",
                            "value": "true"
                        },
                        {
                            "name": "all.notify-output",
                            "value": "false"
                        },
                        {
                            "name": "all.debug-rows",
                            "value": "20"
                        },
                        {
                            "name": "dataflow.master",
                            "value": "yarn"
                        },
                        {
                            "name": "dataflow.deploy-mode",
                            "value": "client"
                        },
                        {
                            "name": "dataflow.queue",
                            "value": "merce.normal"
                        },
                        {
                            "name": "dataflow.num-executors",
                            "value": "2"
                        },
                        {
                            "name": "dataflow.driver-memory",
                            "value": "512M"
                        },
                        {
                            "name": "dataflow.executor-memory",
                            "value": "1G"
                        },
                        {
                            "name": "dataflow.executor-cores",
                            "value": "2"
                        },
                        {
                            "name": "dataflow.verbose",
                            "value": "true"
                        },
                        {
                            "name": "dataflow.local-dirs",
                            "value": ""
                        },
                        {
                            "name": "dataflow.sink.concat-files",
                            "value": "true"
                        }
                    ],
                    "startTime": get_time()
                },
                "flowId": flow_id,
                "flowName": flow_name,
                "flowType": flow_type,
                "name": flow_name + 'scheduler' + str(random.randint(0, 9999))+str(random.randint(0, 9999)),
                "schedulerId": "once",
                "source": "rhinos"
            }
            data_list.append(data)
        print("------���ش��������data------")
        # print(data_list)
        return data_list

    def create_new_scheduler(self):
        """
        ��������scheduler��
        ������scheduler_id_list�� ��get_execution_info(self)����
        :return: scheduler_id_list
        """
        print("------��ʼ��������------")
        from basic_info.url_info import create_scheduler_url
        scheduler_id_list = []
        scheduler_number = 1
        for data in self.data_for_create_scheduler():
            res = requests.post(url=create_scheduler_url, headers=get_headers(host), json=data)
            print('��%d ��scheduler%s' % (scheduler_number,res.text))
            scheduler_number += 1
            time.sleep(2)
            # print(res.status_code, res.text)
            if str(res.status_code).startswith('2') and res.text:
                scheduler_id_format = dict_res(res.text)
                try:
                    scheduler_id = scheduler_id_format["id"]
                except KeyError as e:
                    print("scheduler_id_format�д����쳣%s" % e)
                else:
                    scheduler_id_list.append(scheduler_id)
            else:
                print("flow: %s scheduler����ʧ��" % data["flowid"])
                # return None
        print("------create_new_scheduler(self)ִ�н���, ����scheduler_id_list------\n")
        print('scheduler_id_list', scheduler_id_list)
        return scheduler_id_list

    def get_execution_info(self):
        """
        ����schedulers id ��ѯ��execution id, name, status��flow_id,����������Ϣ��list����ʽ�洢��
        ����scheduler���ѯexecution���ӳ٣���Ҫ�ӵȴ�ʱ��
        :return: e_info_list
        """
        print("------��ѯexecution��Ϣ------")
        scheduler_id_list = self.create_new_scheduler()
        if scheduler_id_list:
            e_info_list = []
            count = 1
            for scheduler_id in scheduler_id_list:
                print('��ʼ��%d �β�ѯ����ѯscheduler id Ϊ %s ��execution info' % (count, scheduler_id))
                # �ȴ�30S���ѯ
                time.sleep(20)
                # ��û�в鵽execution id�� ��Ҫ�ٴβ�ѯ
                e_info = self.get_e_finial_status(scheduler_id)
                e_info_list.append(e_info)
                print('e_info_list:', e_info_list)
                count += 1
            # print('��ѯ�õ���e_info_list', e_info_list)
            print("------��ѯ execution��Ϣ����------\n")
            return e_info_list
        else:
            print("���ص�scheduler_id_listΪ��", scheduler_id_list)
            return None

    def get_e_finial_status(self, scheduler_id):
        """
        ����get_execution_info(self)���ص�scheduler id,
        ��ѯ��scheduler��execution id, name, status,flow id
        :return: e_info
        """
        print("------��ѯexecution����ִ��״̬------")
        if scheduler_id:
            execution_sql = 'select id, status_type, flow_id , flow_scheduler_id from merce_flow_execution where flow_scheduler_id = "%s" ' % scheduler_id
            time.sleep(10)
            select_result = self.ms.ExecuQuery(execution_sql)
            if select_result:
                e_info = {}
                # �Ӳ�ѯ�����ȡֵ
                try:
                    e_info["e_id"] = select_result[0]["id"]
                    e_info["flow_id"] = select_result[0]["flow_id"]
                    e_info["flow_scheduler_id"] = select_result[0]["flow_scheduler_id"]
                    e_info["e_final_status"] = select_result[0]["status_type"]
                except IndexError as e:
                    print("ȡֵʱ���� %s" % e)
                    raise e
                print("------��ѯexecution����ִ��״̬����������execution id��status------\n")
                return e_info
            else:
                return
        else:
            return

    def get_execution_out_put(self):
        """
        ���շ���executionִ�гɹ����dataset id
        :return: dataset id
        """
        print("------��ʼ��ѯexecution�����------")
        e_info_list = self.get_execution_info()
        print("���ص�e_info: %s " % e_info_list)
        # ������ȱʧ�����к������ж�
        if len(e_info_list) == len(self.get_flow_id()):  # flow id ������ͬ
            sink_dataset_list = []
            for i in range(len(e_info_list)):
                sink_dataset_dict = {}
                try:
                    e_id = e_info_list[i]["e_id"]
                    e_final_status = e_info_list[i]["e_final_status"]
                    e_scheduler_id = e_info_list[i]["flow_scheduler_id"]
                    e_flow_id = e_info_list[i]["flow_id"]
                except Exception as e:
                    print('��ȷ�ϸ�flow��e_info_list:')
                else:
                    sink_dataset_dict["flow_id"] = e_flow_id
                    sink_dataset_dict["execution_id"] = e_id
                    sink_dataset_dict["flow_scheduler_id"] = e_scheduler_id
                    if e_id:
                        print("------��ʼ��%s ����״̬���ж�------\n" % e_id)
                        while e_final_status in ("READY", "RUNNING"):
                            print("------����whileѭ��------\n")
                            # ״̬Ϊ ready ���� RUNNINGʱ���ٴβ�ѯe_final_status
                            print("------��ѯǰ�ȴ�5S------\n")
                            time.sleep(5)
                            # ����get_e_finial_status(e_scheduler_id)�ٴβ�ѯ״̬
                            e_info = self.get_e_finial_status(e_scheduler_id)
                            # ��e_final_status ���¸�ֵ
                            e_final_status = e_info["e_final_status"]
                            print("------�ٴβ�ѯ���e_final_status: %s------\n" % e_final_status)
                            # time.sleep(50)
                        if e_final_status in ("FAILED", "KILLED"):
                            print("execution %s ִ��ʧ��" % e_id)
                            sink_dataset_dict["e_final_status"] = e_final_status
                            sink_dataset_dict["o_dataset"] = ""
                            sink_dataset_list.append(sink_dataset_dict)
                            # continue
                        elif e_final_status == "SUCCEEDED":
                            # �ɹ����ѯflow_execution_output���е�dataset, ��sink��Ӧ�����dataset��ȡ��dataset id �����ظ�ID����������Ԥ���ӿڲ鿴����
                            sink_dataset_dict["e_final_status"] = e_final_status
                            # print(e_final_status, e_id)
                            data_json_sql = 'select b.dataset_json from merce_flow_execution as a  LEFT JOIN merce_flow_execution_output as b on a.id = b.execution_id where a.id ="%s"' % e_id
                            print(data_json_sql)
                            data_json = self.ms.ExecuQuery(data_json_sql)
                            print("��ӡdata_json:", data_json)
                            for n in range(len(data_json)):
                                sink_dataset = data_json[n]["dataset_json"]  # ���ؽ��ΪԪ��
                                print('-----sink_dataset-----', sink_dataset)
                                if sink_dataset:
                                    sink_dataset_id = dict_res(sink_dataset)["id"]  # ȡ��json���е�dataset id
                                    sink_dataset_dict["dataset_id"] = sink_dataset_id
                                    d = json.loads(json.dumps(sink_dataset_dict))
                                    sink_dataset_list.append(d)
                                else:
                                    continue
                        else:
                            print("���ص�execution ִ��״̬����ȷ")
                            return
                    else:
                        print("execution������")
                        return
            return sink_dataset_list
        else:
            print("���ص�scheduler_id_listֵȱʧ")
            return

    def get_json(self):
        from util.count_items import count_items
        """
        1.ͨ��datasetԤ���ӿ�ȡ�����ݵ�Ԥ��json�� result.text
        2.�ж�dataset id�Ƿ���ڣ�������ͨ�����ݼ�Ԥ���ӿ�ȡ��Ԥ��������ҵ�������ȵ�dataset id��д��Ԥ�����
        3.Nokia��������ʹ�õ��Ǿɰ��Ԥ���ӿڣ�0.8.11���Ժ�߰汾����ʹ�õ�����Ԥ���ӿ�
        4.flow id ���ʱ����execution id ,yarn�ϵ�log URL ��ִ��״̬д�����
        5.�ԱȽ��
        ��ע������sink multiouput flow���⴦�����а�flow idȫ��д�룬����dataset������ִ�к���д��
        :return:
        """
        print("------��ʼִ��get_json()------")
        sink_output_info = self.get_execution_out_put()
        flow_table = load_workbook(abs_dir(self.table))
        flow_sheet = flow_table.get_sheet_by_name(self.table_sheet)
        sheet_rows = flow_sheet.max_row  # ��ȡ����
        print("��ӡsink_output_info�Ľ��%s"%sink_output_info)
        flow_id_list = self.get_flow_id_not_distinct()  # ��ȡflow����δȥ�ص�����flow id
        sink_multi_flow_id = count_items(flow_id_list)[0]  # �ظ���flow id
        sink_multi_flow_sink_num = count_items(flow_id_list)[1]  # �ظ���flow id ���ظ��Ĵ���
        # ����multi_sink
        if sink_multi_flow_id:
            multi_flow_id = sink_multi_flow_id[0]
            multi_flow_sink_num = sink_multi_flow_sink_num[0]
            for j in range(2, sheet_rows + 1):
                    if multi_flow_id == flow_sheet.cell(row=j, column=2).value:
                        for i in range(0, len(sink_output_info)):
                            log_url = GetLog(sink_output_info[i]["execution_id"], self.host).get_log_url()
                            dataset_id = sink_output_info[i]["dataset_id"]
                            flow_id = sink_output_info[i]["flow_id"]
                            if flow_id == multi_flow_id:
                                if '57' in self.host:
                                    priview_url = "%s/api/datasets/%s/preview?rows=5000&tenant=%s&rows=50" % (
                                        self.host, dataset_id, get_tenant(self.host))
                                    res = requests.get(url=priview_url, headers=get_headers(self.host))
                                    result = res.text
                                else:
                                    statementID = statementId_flow_use(self.host, dataset_id)
                                    result = preview_result_flow_use(self.host, dataset_id,statementID)
                                flow_sheet.cell(row=j+i, column=4, value=dataset_id)
                                flow_sheet.cell(row=j+i, column=8, value=str(result))


        for i in range(0, len(sink_output_info)):
            try:
                dataset_id = sink_output_info[i]["dataset_id"]
                flow_id = sink_output_info[i]["flow_id"]
            except:
                print('��ȷ��sink_output_info�Ľ��:')
            else:
                if '57' in self.host:
                    priview_url = "%s/api/datasets/%s/preview?rows=5000&tenant=%s&rows=50" % (
                        self.host, dataset_id, get_tenant(self.host))
                    res = requests.get(url=priview_url, headers=get_headers(self.host))
                    result = res.text
                else:
                    statementID = statementId_flow_use(self.host, dataset_id)
                    result = preview_result_flow_use(self.host, dataset_id, statementID)
                    print("��ӡԤ�����ݼ��Ľ��%s"%result)
                for j in range(2, sheet_rows + 1):  # ������������ѭ��
                        log_url = GetLog(sink_output_info[i]["execution_id"], self.host).get_log_url()
                        # ��� dataset id��Ⱦ�д��ʵ�ʽ��������Ⱦ�������
                        if dataset_id == flow_sheet.cell(row=j, column=4).value:
                            flow_sheet.cell(row=j, column=8, value=str(result))  # dataset id ��ȣ�ʵ�ʽ��д����
                        # flow id ���ʱ����execution id ,yarn�ϵ�log URL ��ִ��״̬д��
                        if flow_id == flow_sheet.cell(row=j, column=2).value:
                            # print(sink_dataset[i]["flow_id"])
                            flow_sheet.cell(row=j, column=5, value=sink_output_info[i]["execution_id"])
                            flow_sheet.cell(row=j, column=12, value=log_url)
                            flow_sheet.cell(row=j, column=6, value=sink_output_info[i]["e_final_status"])
                        else:
                            for t in range(j, 2, -1):
                                if flow_id == flow_sheet.cell(row=t - 1, column=2).value:
                                    flow_sheet.cell(row=j, column=5, value=sink_output_info[i]["execution_id"])
                                    flow_sheet.cell(row=j, column=12, value=log_url)
                                    flow_sheet.cell(row=j, column=6, value=sink_output_info[i]["e_final_status"])
                                    # flow_sheet.cell(row=j, column=8, value=result.text)  # ʵ�ʽ��д����
                                    break
        flow_table.save(abs_dir(self.table))

        table = load_workbook(abs_dir(self.table))
        table_sheet = table.get_sheet_by_name(self.table_sheet)
        c_rows = table_sheet.max_row
        print('-----��ʼ�ԱȽ��----')
        for i in range(2, c_rows+1):
            table_sheet.cell(row=i, column=1, value=i-1)
            # ��sample step�漰��flow�������н���ж�
            if table_sheet.cell(row=i, column=2).value == '0822a0a2-ce58-42cb-82de-f2ec434b5d94':  #
                if table_sheet.cell(row=i, column=6).value == 'SUCCEEDED' and table_sheet.cell(row=i, column=8):
                    new_result = []
                    expect_result = list(eval(table_sheet.cell(row=i, column=7).value))
                    actual_result = list(eval(table_sheet.cell(row=i, column=8).value))
                    for item in actual_result:
                        if item in expect_result:
                            new_result.append(item)
                    if new_result == actual_result:
                        table_sheet.cell(row=i, column=9, value="pass")
                        # print('test_result:', table_sheet.cell(row=i, column=9).value)
                        table_sheet.cell(row=i, column=10, value="")
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="execution: %s Ԥ�ڽ��ʵ�ʽ����һ�� " %
                                                                 (table_sheet.cell(row=i, column=5).value))
                elif table_sheet.cell(row=i, column=6).value == 'SUCCEEDED' and table_sheet.cell(row=i, column=8) == "":
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="execution: %s Ԥ�ڽ��ʵ�ʽ����һ��,ʵ�ʽ��Ϊ�� " %
                                                             (table_sheet.cell(row=i, column=5).value))
                elif table_sheet.cell(row=i, column=6).value == 'FAILED'or table_sheet.cell(row=i, column=6).value == '':
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="execution: %s ִ��״̬Ϊ %s" % (
                        table_sheet.cell(row=i, column=5).value, table_sheet.cell(row=i, column=6).value))
                else:
                    print('��ȷ��flow_id: %s��ִ��״̬' % table_sheet.cell(row=i, column=2).value)

            else:
                # �ж�modeΪoverwrite
                if table_sheet.cell(row=i, column=11).value == 'overwrite':  # �ж�mode
                    # ʵ�ʽ�����ڲ���ִ�н��Ϊsucceeded
                    if table_sheet.cell(row=i, column=8).value and table_sheet.cell(row=i, column=6).value == "SUCCEEDED":
                        # va7ΪԤ�ڽ����va8Ϊʵ�ʽ�����������������Ա��Ƿ����
                        va7 = list(eval(table_sheet.cell(row=i, column=7).value))
                        va8=table_sheet.cell(row=i, column=8).value
                          # ע�⣺�ǲ�����Ҫ����if �����棿������
                        if va7 != [] and eval(table_sheet.cell(row=i, column=8).value).__class__ == [].__class__ :
                            va8 = list(eval(table_sheet.cell(row=i, column=8).value))
                            va7_k = va7[0].keys()
                            va7_key = list(va7_k)
                            # ���Ų�ͬ��key��������ֻҪ������һ��key�������ȣ�����Ϊ����������
                            result = []
                            for t in range(len(va7_key)):
                                S_va7 = sorted(va7, key=lambda item: item[va7_key[t]], reverse=True)  # û�� idʱ�������
                                S_va8 = sorted(va8, key=lambda item: item[va7_key[t]], reverse=True)
                                result.append(S_va7 == S_va8)
                            print('-----ȷ�Ͻ��------')
                            if True in result:
                                table_sheet.cell(row=i, column=9, value="pass")
                                print('test_result:', table_sheet.cell(row=i, column=9).value)
                                table_sheet.cell(row=i, column=10, value="")
                            else:
                                table_sheet.cell(row=i, column=9, value="fail")
                                table_sheet.cell(row=i, column=10, value="flowname: %s --->Ԥ�ڽ��ʵ�ʽ����һ�� \n" %
                                                                         (table_sheet.cell(row=i, column=3).value))
                        elif va7 == [] and list(eval(table_sheet.cell(row=i, column=8).value)) == []:
                            table_sheet.cell(row=i, column=9, value="pass")
                            # print('test_result:', table_sheet.cell(row=i, column=9).value)
                            table_sheet.cell(row=i, column=10, value="")
                        else:
                            table_sheet.cell(row=i, column=9, value="")
                            table_sheet.cell(row=i, column=10, value="��ȷ��Ԥ�ڽ����ʵ�ʽ��")
                    elif table_sheet.cell(row=i, column=6).value == "FAILED":
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="flowname: %s --->ִ��״̬Ϊ %s\n" % (table_sheet.cell(row=i, column=3).value, table_sheet.cell(row=i, column=6).value))
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="����������datasetID��д����")
                # �ж�modeΪappend
                elif table_sheet.cell(row=i, column=11).value == 'append':
                    if table_sheet.cell(row=i, column=8).value and table_sheet.cell(row=i, column=6).value == "SUCCEEDED":  # ʵ�ʽ������
                        expect_result_list = list(eval(table_sheet.cell(row=i, column=7).value))
                        expect_len = len(expect_result_list)
                        actual_result_list = list(eval(table_sheet.cell(row=i, column=8).value))

                        if expect_result_list == actual_result_list[-expect_len:]:  # ʵ�ʽ����Ƭ��Ԥ�ڽ������һ�µ����ݽ���Ƿ����
                            table_sheet.cell(row=i, column=9, value="pass")
                            table_sheet.cell(row=i, column=10, value="")
                        else:
                            table_sheet.cell(row=i, column=9, value="fail")
                            table_sheet.cell(row=i, column=10,
                                                   value="execution: %s Ԥ�ڽ��ʵ�ʽ����һ�� \nԤ�ڽ��: %s\nʵ�ʽ��: %s" % (
                                                   table_sheet.cell(row=i, column=5).value,
                                                   table_sheet.cell(row=i, column=7).value,
                                                   table_sheet.cell(row=i, column=8).value))
                    elif table_sheet.cell(row=i, column=6).value == "FAILED":  # executionִ��ʧ��
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10,
                                               value="execution: %s ִ��״̬Ϊ %s" % (
                                               table_sheet.cell(row=i, column=5).value, table_sheet.cell(row=i, column=6).value))
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="����������datasetID��д����")

                else:
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="��ȷ��flow��mode")
        table.save(abs_dir("flow_dataset_info.xlsx"))
        print('����������')


if __name__ == '__main__':
    GetCheckoutDataSet().get_json()
    # print(GetCheckoutDataSet().get_flow_id())



