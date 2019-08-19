# coding:utf-8
from basic_info.get_auth_token import get_headers
import unittest
import requests
import json
import time
from util.format_res import dict_res, get_time
from util.Open_DB import MYSQL
from basic_info.url_info import *
from util.data_from_db import create_schedulers
import random
from openpyxl import load_workbook


# 配置数据库连接
ms = MYSQL(MySQL_CONFIG["HOST"], MySQL_CONFIG["USER"], MySQL_CONFIG["PASSWORD"], MySQL_CONFIG["DB"])
abs_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))

# 该类用来测试创建scheduler接口
class CreateSchedulers(unittest.TestCase):
    """用来测试创建schedulers"""
    # 创建schedulers的API路径
    def test_case01(self):
        """创建schedulers，单次执行"""
        scheduler_name = 'api_auto_create_schedulers_once' + str(random.randint(0, 99999))
        flow_table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
        info_sheet = flow_table.get_sheet_by_name("flow_info")
        flow_id = info_sheet.cell(row=2, column=2).value
        flow_name = info_sheet.cell(row=2, column=3).value
        data = {"name": scheduler_name,
                "flowId": flow_id,
                "flowName": flow_name,
                "flowType": 'dataflow',
                "schedulerId": "once",
                "configurations":
                    {"startTime": get_time(), "arguments": [], "cron": "once", "properties": []}
                }
        res = requests.post(url=create_scheduler_url, headers=get_headers(), json=data)
        print(res.status_code, res.text)
        self.assertEqual(res.status_code, 201, '创建单次执行的scheduler失败: %s' % res.text)
        time.sleep(5)

    def test_case02(self):
        """创建schedulers，周期执行"""
        scheduler_name = 'api_auto_create_schedulers_cron' + str(random.randint(0, 99999))
        # start_time = get_time()+(600*1000)  # starttime设为当前时间10分钟后
        start_time = get_time()  # starttime设为当前时间
        end_time = get_time() + (24*3600*1000)  # endtime设为当前时间1天后
        flow_table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
        info_sheet = flow_table.get_sheet_by_name("flow_info")
        flow_id = info_sheet.cell(row=2, column=2).value
        flow_name = info_sheet.cell(row=2, column=3).value
        data = {"name": scheduler_name,
                "flowId": flow_id,
                "flowName": flow_name,
                "flowType": 'dataflow',
                "schedulerId": "cron",
                "source": "rhinos",
                "configurations":
                    {"arguments": [],
                     "cron": "0 0 8 * * ? ",
                     "cronType": "simple",
                     "endTime": end_time,
                     "properties":
                         [{"name":"all.debug","value":"false"},
                          {"name":"all.dataset-nullable","value":"false"},
                          {"name":"all.notify-output","value":"false"},
                          {"name":"all.debug-rows","value":"20"},
                          {"name":"dataflow.master","value":"yarn"},
                          {"name":"dataflow.queue","value":["default"]},
                          {"name":"dataflow.num-executors","value":"2"},
                          {"name":"dataflow.driver-memory","value":"512M"},
                          {"name":"dataflow.executor-memory","value":"1G"},
                          {"name":"dataflow.executor-cores","value":"2"},
                          {"name":"dataflow.verbose","value":"true"},
                          {"name":"dataflow.local-dirs","value":""},
                          {"name":"dataflow.sink.concat-files","value":"true"}],
                     "startTime": start_time}}
        res = requests.post(url=create_scheduler_url, headers=get_headers(), json=data)
        print(res.status_code, res.text)
        self.assertEqual(res.status_code, 201, '创建周期执行的scheduler失败:%s' % res.text)
        time.sleep(5)

# 该类用来测试scheduler查询接口
class SelectSchedulers(unittest.TestCase):
    """测试scheduler查询接口"""
    def test_case01(self):
        """用来id查询scheduler"""
        res = requests.get(url=select_by_schedulerId_url, headers=get_headers(HOST_189))
        schedulerid = dict_res(res.text)["id"]
        # print(scheduler_id)
        # print(type(text), text)
        # 响应码
        self.assertEqual(res.status_code, 200, msg='scheduler查询失败')
        # 查询到的scheduler id 和 用来做查询使用的 scheduler id 做比对
        self.assertEqual(schedulerid, scheduler_id, "通过scheduler ID查询返回的scheduler结果不正确")


# 该类用来测试查询schedulers接口 /api/schedulers/query
class QuerySchedulers(unittest.TestCase):
    """测试查询schedulers接口 /api/schedulers/query"""
    from basic_info.url_info import query_scheduler_url
    def test_case01(self):
        """根据scheduler name模糊查询"""
        keyword = "%student%"
        data = {"fieldList": [{"fieldName": "name", "fieldValue": keyword, "comparatorOperator":"LIKE"}],
                "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"},
                "offset": 0,
                "limit": 8
                }
        # 提取出参数中的查询关键词
        fieldValue = data["fieldList"][0]["fieldValue"][1:-1]
        res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        # print(res.status_code, res.text)
        query_results = dict_res(res.text)
        # print(type(query_results["content"]))
        self.assertEqual(res.status_code, 200, "查询失败")
        # 对比查询关键字和查询结果中的scheduler name
        query_results = dict_res(query_results["content"][0])  # 将查询结果中的第一个值进行dictionary格式化
        query_result_name = query_results["name"]

        # 查询关键词应该包含在查询结果的scheduler name中
        self.assertIn(fieldValue, query_result_name, "查询结果中scheduler的name和查询关键词name不一致")

    def test_case02(self):
        """根据flowtype-dataflow查询"""
        data = {"fieldList":
                    [{"fieldName": "flowType", "fieldValue": "dataflow", "comparatorOperator": "LIKE"}],
                "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"}, "offset": 0, "limit": 8
                }
        # 提取出参数中的查询关键词
        fieldValue = data["fieldList"][0]["fieldValue"]

        res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        query_results = dict_res(res.text)
        # print(type(query_results["content"]))
        # 将查询结果中的第一个值进行dictionary格式化
        query_results = dict_res(query_results["content"][0])
        query_result_flowType = query_results["flowType"]
        # 响应码应该为200
        self.assertEqual(res.status_code, 200, "查询失败")
        # 对比查询关键字和查询结果中的flowType
        self.assertEqual(fieldValue, query_result_flowType, "查询结果中scheduler关联flowtype和查询关键词flowType不一致")

    def test_case03(self):
        """根据flowtype-workflow查询"""
        data = {"fieldList":
                    [{"fieldName": "flowType", "fieldValue": "workflow", "comparatorOperator": "LIKE"}],
                "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"}, "offset": 0, "limit": 8
                }
        # 提取出参数中的查询关键词
        fieldValue = data["fieldList"][0]["fieldValue"]
        res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        # query_results = dict_res(res.text)
        # # print(type(query_results["content"]))
        # # 将查询结果中的第一个值进行dictionary格式化
        # query_results = dict_res(query_results["content"][0])
        # query_result_flowType = query_results["flowType"]
        # 响应码应该为200
        self.assertEqual(res.status_code, 200, "查询失败")
        # 对比查询关键字和查询结果中的flowType
        # print("fieldValue", fieldValue, "query_result_flowType", query_result_flowType)
        # self.assertEqual(fieldValue, query_result_flowType, "查询结果中scheduler关联flowtype和查询关键词flowType不一致")

    def test_case04(self):
        """根据flowtype-streamflow查询"""
        data = {"fieldList":
                    [{"fieldName": "flowType", "fieldValue": "streamflow", "comparatorOperator": "LIKE"}],
                "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"}, "offset": 0, "limit": 8
                }
        # 提取出参数中的查询关键词
        fieldValue = data["fieldList"][0]["fieldValue"]

        res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        query_results = dict_res(res.text)
        # print(type(query_results["content"]))
        # 将查询结果中的第一个值进行dictionary格式化
        # query_results = dict_res(query_results["content"][0])
        # query_result_flowType = query_results["flowType"]
        # 响应码应该为200
        self.assertEqual(res.status_code, 200, "查询失败")
        # 对比查询关键字和查询结果中的flowType
        # print("fieldValue", fieldValue, "query_result_flowType", query_result_flowType)
        # self.assertEqual(fieldValue, query_result_flowType, "查询结果中scheduler关联flowtype和查询关键词flowType不一致")

    def test_case05(self):
            """flowtype+name组合查询scheduler"""
            data ={"fieldList":
                       [{"fieldName":"name","fieldValue":"%gbj%","comparatorOperator":"LIKE"},
                        {"fieldName":"flowType","fieldValue":"workflow","comparatorOperator":"EQUAL"}],
                   "sortObject":{"field":"lastModifiedTime","orderDirection":"DESC"},
                   "offset":0,
                   "limit":8}
            data_name = data["fieldList"][0]["fieldValue"][1:-1]
            data_flowType = data["fieldList"][1]["fieldValue"]
            res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
            # print(res.status_code, res.text)
            #
            # query_results = dict_res(res.text)
            # query_result_name = query_results["content"][0]["name"]
            # query_result_flowType = query_results["content"][0]["flowType"]
            # print(data_name, query_result_name)
            self.assertEqual(200, res.status_code, "flowtype+name组合查询scheduler失败:%s" % res.text)
            # self.assertIn(data_name, query_result_name, "查询出的scheduler name 不包含查询关键词")
            # print(data_flowType, query_result_flowType)
            # self.assertEqual(data_flowType, query_result_flowType,  "查询出的scheduler flowType和查询条件不一致")


    def test_case06(self):
        """query:根据上次修改时间查询全部的scheduler"""
        end_time = get_time()  # lastModifiedTime结束时间是当前时间
        start_time = get_time() - (10 * 24 * 3600 * 1000)  # lastModifiedTime开始时间是当前时间的十天前
        data = {"fieldList": [
            {"fieldName": "lastModifiedTime", "fieldValue": start_time, "comparatorOperator": "GREATER_THAN"},
            {"fieldName": "lastModifiedTime", "fieldValue": end_time, "comparatorOperator": "LESS_THAN"}
        ],
            "sortObject": {"field": "lastModifiedTime", "orderDirection": "DESC"},
            "offset": 0,
            "limit": 8
        }
        res = requests.post(url=self.query_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))

        query_results = dict_res(res.text)
        # print(res.text, query_results)
        first_Time = query_results["content"][0]["lastModifiedTime"]
        # print('first_one_lastModifiedTime:', first_Time)
        # 将查询结果中的第一个的lastModifiedTime和查询使用的开始时间，结束时间做对比，应该包含在二者之间
        self.assertEqual(end_time > first_Time > start_time, True,
                         "查询结果的lastModifiedTime不包含在起始时间内，查询结果不正确")


# 该类用来测试启用停用计划接口
class EnableDisable(unittest.TestCase):
    """测试启用停用、批量删除schedulers接口"""
    def test_case01(self):
        """启用计划"""
        data = []
        data.append(scheduler_id)
        res = requests.post(url=enable_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        # print(res.status_code)
        self.assertEqual(res.status_code, 204, msg="启用计划接口调用失败")

    def test_case02(self):
        """停用计划"""
        data = []
        data.append(scheduler_id)
        res = requests.post(url=disable_scheduler_url, headers=get_headers(HOST_189), data=json.dumps(data))
        # print(res.status_code)
        self.assertEqual(res.status_code, 204, msg="停用计划接口调用失败")

    def test_case03(self):
        """批量删除计划"""
        from basic_info.url_info import remove_list_url
        data = []
        id1 = create_schedulers()
        data.append(id1)
        time.sleep(2)
        id2 = create_schedulers()
        data.append(id2)
        res = requests.post(url=remove_list_url, headers=get_headers(HOST_189), json=data)
        self.assertEqual(res.status_code, 200, "批量删除接口调用失败")


# 该类用来测试update schedulers接口  需要开发做修改，暂不测试该类接口
# class UpdateScheduler(unittest.TestCase):
#     """测试update计划接口, update name"""
#     def test_case01(self):
#         """schedulers更新为周期执行"""
#         from basic_info.url_info import update_scheduler_url
#         scheduler_name = time.strftime("%Y%m%d%H%M%S", time.localtime()) + 'update_schedulers'
#         data = {"id": scheduler_id,
#                 "name": scheduler_name,
#                 "schedulerId": "cron",
#                 "flowId": flow_id, "flowName": get_flows()[0][0], "flowType": get_flows()[0][1],
#                 "configurations": {"cron": "0 0 19 * * ? ", "arguments": [], "startTime": get_time(), "endTime": get_time()+(2*24*3600*1000), "properties":[{"name":"all.debug","value":"false"},{"name":"all.dataset-nullable","value":"false"},{"name":"all.lineage.enable","value":"true"},{"name":"all.notify-output","value":"false"},{"name":"all.debug-rows","value":"20"},{"name":"dataflow.master","value":"yarn"},{"name":"dataflow.deploy-mode","value":"client"},{"name":"dataflow.queue","value":"default"},{"name":"dataflow.num-executors","value":"2"},{"name":"dataflow.driver-memory","value":"512M"},{"name":"dataflow.executor-memory","value":"1G"},{"name":"dataflow.executor-cores","value":"2"},{"name":"dataflow.verbose","value":"true"},{"name":"dataflow.local-dirs","value":""},{"name":"dataflow.sink.concat-files","value":"true"}]},
#                 }
#         res = requests.put(url=update_scheduler_url, headers=get_headers(), json=data)
#         # print(res.status_code, res.text)
#         self.assertEqual(res.status_code, 204, '调用更新schedulers的接口失败')


